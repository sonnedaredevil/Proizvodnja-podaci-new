import os
import re

# === START FINALNO CISCENJE IMENA IZ RAZLOGA ===
RECI_KOJE_SE_BRISU_IZ_RAZLOGA = [
    "bojan",
    "dimitrije",
    "smiljanić",
    "smiljanic",
    "miler",
    "jelena",
    "stefanović",
    "stefanovic",
    # Imena kolega koja Excel/notes ponekad ubaci kao razlog zastoja.
    # Brišu se samo iz teksta razloga, a ostatak komentara ostaje.
    "valentina",
    "savatić",
    "savatic",
    "atila",
    "čolak",
    "colak",
    "maja",
    "bukarac",
]

def finalno_ocisti_razlog_od_imena(vrednost):
    """
    Briše samo imena iz razloga/komentara.
    Ostatak teksta ostaje.
    Ako posle brisanja ne ostane ništa korisno, vraća prazan tekst.
    """
    tekst = str(vrednost or "")

    for rec in RECI_KOJE_SE_BRISU_IZ_RAZLOGA:
        tekst = re.sub(
            r"(?<!\w)" + re.escape(rec) + r"(?!\w)",
            " ",
            tekst,
            flags=re.IGNORECASE
        )

    tekst = re.sub(r"\s+", " ", tekst).strip()
    tekst = re.sub(r"^[\.\,\;\:\-\–\—\_\s]+", "", tekst).strip()
    tekst = re.sub(r"[\.\,\;\:\-\–\—\_\s]+$", "", tekst).strip()
    tekst = re.sub(r"\s+", " ", tekst).strip()

    return tekst


def finalno_ocisti_df_razloge(df):
    """
    Čisti kolonu Razlog i slične tekstualne kolone.
    Ne briše ceo notes ako ima ostatak teksta.
    Briše red samo ako je razlog bio samo ime kolege i posle čišćenja ostane prazno.
    """
    if df is None or len(df) == 0:
        return df

    df = df.copy()

    kolone = [
        "Razlog",
        "Komentar",
        "Comment",
        "Note",
        "Notes",
        "Opis",
        "Tekst",
    ]

    for kol in kolone:
        if kol in df.columns:
            df[kol] = df[kol].apply(finalno_ocisti_razlog_od_imena)

    if "Razlog" in df.columns:
        df = df[~df["Razlog"].astype(str).apply(da_li_je_excel_threaded_comment_poruka)].copy()
        df = df[df["Razlog"].astype(str).str.strip() != ""].copy()

    return df
# === END FINALNO CISCENJE IMENA IZ RAZLOGA ===

import shutil
import base64
import tempfile
import hashlib
from datetime import datetime, date
import calendar

import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from openpyxl import load_workbook


# ============================================================
# STREAMLIT PODEŠAVANJA
# ============================================================

st.set_page_config(
    page_title="Proizvodnja dashboard",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Naslov se prikazuje kasnije kroz prilagođeni HTML, centrirano.


# ============================================================
# TV / DARK TEMA
# ============================================================

st.markdown(
    """
    <style>
    .stApp {
        background: linear-gradient(135deg, #0f172a 0%, #111827 45%, #020617 100%);
        color: #e5e7eb;
    }

    section[data-testid="stSidebar"] {
        background-color: #020617;
        border-right: 1px solid #1f2937;
    }

    section[data-testid="stSidebar"] * {
        color: #e5e7eb !important;
    }

    h1, h2, h3, h4 {
        color: #f9fafb !important;
    }

    button[data-baseweb="tab"] {
        font-size: 18px;
        font-weight: 700;
    }

    .block-container {
        padding-top: 1.5rem;
        padding-left: 2rem;
        padding-right: 2rem;
        max-width: 100%;
    }

    [data-testid="stDataFrame"] {
        background-color: #0f172a;
    }

    [data-testid="stMetric"] {
        background-color: #020617;
        border: 1px solid #334155;
        border-radius: 12px;
        padding: 10px;
    }

    [data-testid="stMetric"] * {
        color: #f9fafb !important;
    }

    [data-testid="stMetricLabel"] {
        color: #cbd5e1 !important;
        font-size: 13px !important;
    }

    [data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-size: 24px !important;
        font-weight: 800 !important;
    }

    [data-testid="stMetricDelta"] {
        color: #e5e7eb !important;
    }

    div[data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #111827 !important;
        border: 1px solid #334155 !important;
        border-radius: 16px !important;
    }

    div[data-testid="stVerticalBlockBorderWrapper"] * {
        color: #f9fafb;
    }

    /* FIX: sidebar/filteri moraju da imaju sopstveni vertikalni skrol.
       Ovo radi i na novijim verzijama Streamlit-a gde se unutrašnji
       container promeni, pa ciljamo i sidebar i njegov direktni sadržaj. */
    section[data-testid="stSidebar"] {
        height: 100vh !important;
        max-height: 100vh !important;
        overflow-y: auto !important;
        overflow-x: hidden !important;
    }

    section[data-testid="stSidebar"] > div {
        height: 100vh !important;
        max-height: 100vh !important;
        overflow-y: auto !important;
        overflow-x: hidden !important;
        padding-bottom: 5rem !important;
    }

    section[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        height: 100vh !important;
        max-height: 100vh !important;
        overflow-y: auto !important;
        overflow-x: hidden !important;
        padding-bottom: 5rem !important;
    }

    section[data-testid="stSidebar"]::-webkit-scrollbar,
    section[data-testid="stSidebar"] > div::-webkit-scrollbar {
        width: 10px;
    }

    section[data-testid="stSidebar"]::-webkit-scrollbar-thumb,
    section[data-testid="stSidebar"] > div::-webkit-scrollbar-thumb {
        background: #475569;
        border-radius: 999px;
    }
    </style>
    """,
    unsafe_allow_html=True
)



# ============================================================
# DODATNI CSS - CRNA SLOVA U POLJIMA FILTERA
# ============================================================

st.markdown(
    """
    <style>
    section[data-testid="stSidebar"] input {
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important;
    }
    section[data-testid="stSidebar"] div[data-baseweb="select"] input,
    section[data-testid="stSidebar"] div[data-baseweb="select"] span,
    section[data-testid="stSidebar"] div[data-baseweb="select"] div {
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important;
    }
    section[data-testid="stSidebar"] div[data-baseweb="select"] svg {
        fill: #000000 !important;
    }
    section[data-testid="stSidebar"] div[data-baseweb="tag"] {
        background-color: #e2e8f0 !important;
    }
    section[data-testid="stSidebar"] div[data-baseweb="tag"] span {
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important;
    }
    section[data-testid="stSidebar"] [data-testid="stDateInput"] input {
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)


# ============================================================
# INDUSTRIJSKA POZADINA - CHATGPT IMAGE
# ============================================================

def pronadji_pozadinsku_sliku():
    folder_slika = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "slike"
    )

    osnovni_naziv = "ChatGPT Image Jun 5, 2026, 09_35_24 PM"

    moguce_ekstenzije = [
        ".png",
        ".jpg",
        ".jpeg",
        ".webp"
    ]

    for ekstenzija in moguce_ekstenzije:
        putanja = os.path.join(
            folder_slika,
            osnovni_naziv + ekstenzija
        )

        if os.path.exists(putanja):
            return putanja

    return None


def postavi_pozadinsku_sliku():
    putanja_slike = pronadji_pozadinsku_sliku()

    if putanja_slike is None:
        return

    with open(putanja_slike, "rb") as fajl:
        slika_base64 = base64.b64encode(
            fajl.read()
        ).decode("utf-8")

    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image:
                linear-gradient(
                    rgba(7, 15, 27, 0.58),
                    rgba(7, 15, 27, 0.72)
                ),
                url("data:image/png;base64,{slika_base64}") !important;

            background-size: cover !important;
            background-position: center center !important;
            background-repeat: no-repeat !important;
            background-attachment: fixed !important;
        }}

        .stApp::before {{
            content: "";
            position: fixed;
            inset: 0;
            pointer-events: none;
            background:
                radial-gradient(
                    circle at 75% 15%,
                    rgba(14, 165, 233, 0.10),
                    transparent 36%
                ),
                linear-gradient(
                    180deg,
                    rgba(2, 6, 23, 0.08),
                    rgba(2, 6, 23, 0.20)
                );
            z-index: 0;
        }}

        .stApp > * {{
            position: relative;
            z-index: 1;
        }}

        section[data-testid="stSidebar"] {{
            background: rgba(2, 6, 23, 0.88) !important;
            backdrop-filter: blur(12px);
            -webkit-backdrop-filter: blur(12px);
        }}

        div[data-testid="stVerticalBlockBorderWrapper"] {{
            background: rgba(15, 23, 42, 0.78) !important;
            backdrop-filter: blur(8px);
            -webkit-backdrop-filter: blur(8px);
        }}

        [data-testid="stMetric"] {{
            background: rgba(2, 6, 23, 0.82) !important;
            backdrop-filter: blur(8px);
            -webkit-backdrop-filter: blur(8px);
        }}

        [data-testid="stDataFrame"] {{
            background: rgba(15, 23, 42, 0.86) !important;
            backdrop-filter: blur(7px);
            -webkit-backdrop-filter: blur(7px);
        }}
        </style>
        """,
        unsafe_allow_html=True
    )


postavi_pozadinsku_sliku()


# ============================================================
# FAJLOVI PO MESECIMA
# ============================================================

FOLDER_IZVESTAJA = r"R:\Operations\1. Proizvodnja\13. Dnevni izvestaj proizvodnje\2026"

FAJLOVI_PO_MESECIMA = {
    "Januar 2026": os.path.join(FOLDER_IZVESTAJA, "01. Production realization JANUAR.xlsx"),
    "Februar 2026": os.path.join(FOLDER_IZVESTAJA, "02. Production realization FEBRUAR.xlsx"),
    "Mart 2026": os.path.join(FOLDER_IZVESTAJA, "03. Production realization MART.xlsx"),
    "April 2026": os.path.join(FOLDER_IZVESTAJA, "04. Production realization APRIL.xlsx"),
    "Maj 2026": os.path.join(FOLDER_IZVESTAJA, "05. Production realization MAJ.xlsx"),
    "Jun 2026": os.path.join(FOLDER_IZVESTAJA, "06. Production realization JUN.xlsx"),
    "Jul 2026": os.path.join(FOLDER_IZVESTAJA, "07. Production realization JUL.xlsx"),
    "Avgust 2026": os.path.join(FOLDER_IZVESTAJA, "08. Production realization AVGUST.xlsx"),
    "Septembar 2026": os.path.join(FOLDER_IZVESTAJA, "09. Production realization SEPTEMBAR.xlsx"),
    "Oktobar 2026": os.path.join(FOLDER_IZVESTAJA, "10. Production realization OKTOBAR.xlsx"),
    "Novembar 2026": os.path.join(FOLDER_IZVESTAJA, "11. Production realization NOVEMBAR.xlsx"),
    "Decembar 2026": os.path.join(FOLDER_IZVESTAJA, "12. Production realization DECEMBAR.xlsx"),
}


# ============================================================
# PODEŠAVANJA ČITANJA
# ============================================================

prvi_red_podataka = 8
poslednji_red_podataka = 40

red_smena = 6
red_nazivi = 7
kolona_datum = 1

dozvoljene_smene = {
    "I shift",
    "II shift",
    "III shift"
}

preskoci_tabove = [
    "AIDA 2 APP350",
    "Realizacija CW 21",
    "Realizacija CW 20",
    "Realizacija CW 22",
    "REALIZACIJA MESECNA",
    "Plan rada linija",
    "Realizacija smena",
    "Analiza skarta cw 18",
    "Analiza skarta cw 19",
    "Analiza skarta cw20",
    "Analiza skarta cw13",
    "Stops analysis CW7",
    "Sheet1",
    "Annealing"
]

iskljuci_iz_ukupnog_proracuna = [
    "AIDA VITESKO",
    "HEATING VITESCO STATOR",
    "HEATING VITESCO ROTOR",
    "DMC EMR4 STATOR",
    "DMC EMR4 ROTOR",
]

dmc_stator_iz_smena_tabovi = [
    "DMC L1 APP550",
    "DMC L1 GP12",
    "DMC L2 APP550",
    "DMC L2 GP12",
    "DMC L3 APP550",
    "DMC L3 GP12",
    "DMC L4 APP550",
    "DMC L4 GP12",
    "DMC L5 APP550",
    "DMC L2 APP350",
    "DMC L3 APP350",
    "DMC L4 APP350",
    "DMC L1 LK-4",
    "Stator L1",
    "Stator L2",
    "Stator L3",
    "Stator L4",
    "Stator L5",
]

rotor_iz_smena_tabovi = [
    "DMC EMR4 ROTOR",
]

KPI_PROCESI = ["STAMPING", "WELDING", "DMC", "ROTOR"]


# ============================================================
# MAPA PROJEKATA
# ============================================================

MAPA_PROJEKATA = {
    "AIDA L1 APP350": "APP350",
    "AIDA 1 APP350": "APP350",
    "AIDA 3 APP350": "APP350",
    "AIDA 8 APP350": "APP350",
    "AIDA 9 APP350": "APP350",
    "DMC L2 APP350": "APP350",
    "DMC L3 APP350": "APP350",
    "DMC L4 APP350": "APP350",
    "Rotor L1 APP350": "APP350",
    "Rotor L3 APP350": "APP350",
    "Rotor L8 APP350": "APP350",

    "AIDA L2": "APP550",
    "AIDA 2": "APP550",
    "AIDA L4": "APP550",
    "AIDA 4": "APP550",
    "AIDA L5": "APP550",
    "AIDA 5": "APP550",
    "AIDA L6": "APP550",
    "AIDA 6": "APP550",
    "AIDA L7": "APP550",
    "AIDA 7": "APP550",
    "Stator L1": "APP550",
    "Stator L2": "APP550",
    "Stator L3": "APP550",
    "Stator L4": "APP550",
    "Stator L5": "APP550",
    "DMC L1 APP550": "APP550",
    "DMC L1 GP12": "APP550",
    "DMC L2 APP550": "APP550",
    "DMC L2 GP12": "APP550",
    "DMC L3 APP550": "APP550",
    "DMC L3 GP12": "APP550",
    "DMC L4 APP550": "APP550",
    "DMC L4 GP12": "APP550",
    "DMC L5 APP550": "APP550",
    "Rotor L2 APP550": "APP550",
    "Rotor L4 APP550": "APP550",
    "Rotor L5 APP550": "APP550",
    "Rotor L6 APP550": "APP550",

    "DMC EMR4 ROTOR": "VITESKO EMR4",
    "DMC EMR4 STATOR": "VITESKO EMR4",
    "HEATING VITESCO ROTOR": "VITESKO EMR4",
    "HEATING VITESCO STATOR": "VITESKO EMR4",
    "AIDA VITESKO": "VITESKO EMR4",

    "AIDA LK-4": "LK-4",
    "AIDA LK4": "LK-4",
    "AIDA LK 4": "LK-4",
    "AIDA LK-3": "LK-4",
    "DMC L1 LK-4": "LK-4",
    "DMC L1 LK4": "LK-4",
    "DMC L1 LK 4": "LK-4",
    "Rotor L10 LK-4": "LK-4",
    "Rotor L10 LK4": "LK-4",
    "Rotor L10 LK 4": "LK-4",

    # Varijante koje se mogu pojaviti u nazivima tabova.
    "AIDA 7 APP350": "APP350",
    "AIDA L7 APP350": "APP350",
    "AIDA 7 APP550": "APP550",
    "AIDA L7 APP550": "APP550",
}


# ============================================================
# MAPA PROCESA
# ============================================================

MAPA_PROCESA = {
    "AIDA L1 APP350": "STAMPING",
    "AIDA 1 APP350": "STAMPING",
    "AIDA L2": "STAMPING",
    "AIDA 2": "STAMPING",
    "AIDA 3 APP350": "STAMPING",
    "AIDA L4": "STAMPING",
    "AIDA 4": "STAMPING",
    "AIDA L5": "STAMPING",
    "AIDA 5": "STAMPING",
    "AIDA L6": "STAMPING",
    "AIDA 6": "STAMPING",
    "AIDA L7": "STAMPING",
    "AIDA 7": "STAMPING",
    "AIDA 8 APP350": "STAMPING",
    "AIDA 9 APP350": "STAMPING",
    "AIDA LK-4": "STAMPING",
    "AIDA LK4": "STAMPING",
    "AIDA LK 4": "STAMPING",
    "AIDA LK-3": "STAMPING",
    "AIDA VITESKO": "STAMPING",
    "AIDA 7 APP350": "STAMPING",
    "AIDA L7 APP350": "STAMPING",
    "AIDA 7 APP550": "STAMPING",
    "AIDA L7 APP550": "STAMPING",

    "Stator L1": "WELDING",
    "Stator L2": "WELDING",
    "Stator L3": "WELDING",
    "Stator L4": "WELDING",
    "Stator L5": "WELDING",
    "HEATING VITESCO STATOR": "WELDING",
    "HEATING VITESCO ROTOR": "WELDING",

    "DMC L1 APP550": "DMC",
    "DMC L2 APP550": "DMC",
    "DMC L3 APP550": "DMC",
    "DMC L4 APP550": "DMC",
    "DMC L5 APP550": "DMC",
    "DMC L2 APP350": "DMC",
    "DMC L3 APP350": "DMC",
    "DMC L4 APP350": "DMC",
    "DMC L1 LK-4": "DMC",
    "DMC L1 LK4": "DMC",
    "DMC L1 LK 4": "DMC",
    "DMC EMR4 STATOR": "DMC",
    "DMC EMR4 ROTOR": "DMC",

    "DMC L1 GP12": "GP12",
    "DMC L2 GP12": "GP12",
    "DMC L3 GP12": "GP12",
    "DMC L4 GP12": "GP12",

    "Rotor L1 APP350": "ROTOR",
    "Rotor L2 APP550": "ROTOR",
    "Rotor L3 APP350": "ROTOR",
    "Rotor L4 APP550": "ROTOR",
    "Rotor L5 APP550": "ROTOR",
    "Rotor L6 APP550": "ROTOR",
    "Rotor L8 APP350": "ROTOR",
    "Rotor L10 LK-4": "ROTOR",
    "Rotor L10 LK4": "ROTOR",
    "Rotor L10 LK 4": "ROTOR",
}


# ============================================================
# MAPIRANJE DNEVNIH KOLONA
# ============================================================

MAPIRANJE = {
    "Plan_STATOR": [
        "Plan per DAY STATOR",
        "Plan DAY STATOR",
        "Plan stator",
        "Plan STATOR",
    ],
    "Plan_ROTOR": [
        "Plan per DAY ROTOR",
        "Plan DAY ROTOR",
        "Plan rotor",
        "Plan ROTOR",
    ],
    "Realizacija_STATOR": [
        "TOTAL Realization STATOR",
        "Total Realization STATOR",
        "Total realization STATOR",
        "TOTAL production STATOR",
        "Total production STATOR",
    ],
    "Realizacija_ROTOR": [
        "Total realization ROTOR",
        "TOTAL Realization ROTOR",
        "Total production ROTOR",
        "TOTAL production ROTOR",
    ],
    "Total_production_AB": [
        "Total production AB",
    ],
    "Total_production_AC": [
        "Total production AC",
    ],
    "Total_production_AD": [
        "Total production AD",
    ],
    "Total_production_AE": [
        "Total production AE",
    ],
    "OK_STATOR": [
        "TOTAL OK parts STATOR",
        "Total OK parts STATOR",
        "OK STATOR",
        "TOTAL Good Parts/stator",
        "Total Good Parts/stator",
        "Good Parts/stator",
        "TOTAL Good Parts / stator",
        "Total Good Parts / stator",
        "Good Parts / stator",
    ],
    "OK_ROTOR": [
        "TOTAL OK parts ROTOR",
        "Total OK parts ROTOR",
        "OK ROTOR",
        "TOTAL Good Parts/rotor",
        "Total Good Parts/rotor",
        "Good Parts/rotor",
        "TOTAL Good Parts / rotor",
        "Total Good Parts / rotor",
        "Good Parts / rotor",
    ],
    "NOK_STATOR": [
        "TOTAL NOK parts STATOR",
        "Total NOK parts STATOR",
        "NOK STATOR",
        "TOTAL NOK Parts/stator",
        "Total NOK Parts/stator",
        "NOK Parts/stator",
        "TOTAL NOK Parts / stator",
        "Total NOK Parts / stator",
        "NOK Parts / stator",
    ],
    "NOK_ROTOR": [
        "TOTAL NOK parts rotor",
        "TOTAL NOK parts ROTOR",
        "Total NOK parts rotor",
        "Total NOK parts ROTOR",
        "NOK ROTOR",
        "TOTAL NOK Parts/rotor",
        "Total NOK Parts/rotor",
        "NOK Parts/rotor",
        "TOTAL NOK Parts / rotor",
        "Total NOK Parts / rotor",
        "NOK Parts / rotor",
    ],
    "Reason": [
        "Reason",
    ],
}


# ============================================================
# OSNOVNE FUNKCIJE
# ============================================================

def ocisti_tekst(vrednost):
    if vrednost is None:
        return None

    tekst = str(vrednost).replace("\n", " ").strip()

    if tekst == "":
        return None

    return tekst



def da_li_je_excel_threaded_comment_poruka(tekst):
    """
    Excel ponekad u tekst komentara ubaci sistemsku poruku o threaded comment-u.
    To nije razlog zastoja i ne sme da uđe u grafikone/tabele.
    """
    t = str(tekst or "").lower()
    obrasci = [
        "threaded comment",
        "your version of excel allows you to read this threaded comment",
        "any edits to it will get removed",
        "go.microsoft.com/fwlink",
        "learn more https",
        "[threaded comment]",
    ]
    return any(o in t for o in obrasci)

def normalizuj_razlog(razlog):
    if da_li_je_excel_threaded_comment_poruka(razlog):
        return "nije upisan razlog"

    # === CISCENJE IMENA U normalizuj_razlog ===
    razlog = finalno_ocisti_razlog_od_imena(razlog)
    if razlog is None:
        return "nije upisan razlog"

    tekst = str(razlog).strip().lower()
    tekst = tekst.replace(":", " ")
    tekst = tekst.strip(" -–—:;,.")
    tekst = " ".join(tekst.split())

    if tekst == "":
        return "nije upisan razlog"

    tekst = re.sub(
        r"otis(?:ak|am|ci)",
        "otisak",
        tekst,
        flags=re.IGNORECASE
    )

    return tekst
def ocisti_note(note):
    if note is None:
        return None

    tekst = str(note)

    ignorisi_pojmove = [
        "Bojan Smiljanić",
        "Bojan Smiljanic",
        "Dimitrije Miler",
        "Jelena Stefanović",
        "Jelena Stefanovic",
        "Valentina Savatić",
        "Valentina Savatic",
        "Atila Čolak",
        "Atila Colak",
        "Maja Bukarac",
        "operacije",
        "Operacije",
        "OPERACIJE",
    ]

    for pojam in ignorisi_pojmove:
        tekst = re.sub(re.escape(pojam), "", tekst, flags=re.IGNORECASE)

    tekst = tekst.replace(":", " ")
    tekst = tekst.replace("\r", "\n").strip()

    linije = []
    for linija in tekst.split("\n"):
        linija = linija.strip()
        if da_li_je_excel_threaded_comment_poruka(linija):
            continue
        if linija:
            linije.append(linija)

    tekst = "\n".join(linije).strip()

    if tekst == "":
        return None

    return tekst


def norm(vrednost):
    tekst = ocisti_tekst(vrednost)

    if tekst is None:
        return ""

    tekst = tekst.lower()
    tekst = " ".join(tekst.split())

    return tekst


def broj(vrednost):
    if vrednost is None:
        return 0

    try:
        return float(vrednost)
    except Exception:
        return 0


def procenat(deo, ukupno):
    if ukupno is None or ukupno == 0:
        return None

    return deo / ukupno * 100


def format_broj(x):
    if x is None or pd.isna(x):
        return "-"

    return f"{x:,.0f}".replace(",", ".")


def format_proc(x):
    if x is None or pd.isna(x):
        return "-"

    return f"{x:.1f}%"


def pojednostavi_naziv_za_mapiranje(vrednost):
    """
    Pravi pomoćni naziv za poređenje naziva tabova/mašina.
    Cilj je da AIDA LK4, AIDA LK-4 i aida lk 4 budu prepoznati isto.
    Takođe izjednačava č/ć/c za nazive koji se unose različito.
    """
    tekst = norm(vrednost)
    tekst = (
        tekst
        .replace("č", "c")
        .replace("ć", "c")
        .replace("š", "s")
        .replace("đ", "dj")
        .replace("ž", "z")
    )
    tekst = re.sub(r"[^a-z0-9]+", "", tekst)
    return tekst


def pronadji_u_mapi(naziv, mapa, default):
    if naziv is None:
        return default

    naziv_cist = str(naziv).strip()

    if naziv_cist in mapa:
        return mapa[naziv_cist]

    naziv_upper = naziv_cist.upper()

    for kljuc, vrednost in mapa.items():
        if kljuc.upper() == naziv_upper:
            return vrednost

    # Dodatno fleksibilno poređenje:
    # ignoriše razliku u crticama, razmacima, velikim/malim slovima i č/c varijante.
    naziv_simple = pojednostavi_naziv_za_mapiranje(naziv_cist)

    for kljuc, vrednost in mapa.items():
        if pojednostavi_naziv_za_mapiranje(kljuc) == naziv_simple:
            return vrednost

    return default


def projekat_iz_masine(masina):
    projekat = pronadji_u_mapi(masina, MAPA_PROJEKATA, "NEMAPIRANO")

    if projekat != "NEMAPIRANO":
        return projekat

    # Rezervna logika ako je naziv taba malo drugačiji od ručno upisane mape,
    # npr. "AIDA 7 APP350", "AIDA LK4", "DMC L1 LK4".
    t = pojednostavi_naziv_za_mapiranje(masina)

    if "app350" in t:
        return "APP350"

    if "app550" in t:
        return "APP550"

    if "lk4" in t or "lk3" in t:
        return "LK-4"

    if "emr4" in t or "vitesco" in t or "vitesko" in t:
        return "VITESKO EMR4"

    return projekat


def proces_iz_masine(masina):
    proces = pronadji_u_mapi(masina, MAPA_PROCESA, "NEMAPIRANO")

    if proces != "NEMAPIRANO":
        return proces

    # Rezervna logika za nazive tabova koji nisu ručno navedeni u mapi.
    t = pojednostavi_naziv_za_mapiranje(masina)

    if t.startswith("aida"):
        return "STAMPING"

    if t.startswith("stator") or "heatingvitescostator" in t or "heatingviteskostator" in t:
        return "WELDING"

    if t.startswith("rotor"):
        return "ROTOR"

    if t.startswith("dmc"):
        if "gp12" in t:
            return "GP12"
        return "DMC"

    if "heatingvitescorotor" in t or "heatingviteskorotor" in t:
        return "WELDING"

    return proces


def tip_proizvoda_iz_proizvoda(proizvod):
    t = str(proizvod).upper()

    if "STATOR" in t:
        return "STATOR"

    if "ROTOR" in t:
        return "ROTOR"

    return "NEPOZNATO"


def slika_u_base64(putanja):
    if not os.path.exists(putanja):
        return None

    with open(putanja, "rb") as f:
        return base64.b64encode(f.read()).decode()


def prikazi_overlay_sliku(putanja_slike, naslov="Slika"):
    img64 = slika_u_base64(putanja_slike)

    if img64 is None:
        st.error(f"Slika nije pronađena: {putanja_slike}")
        return

    st.markdown(
        f"""
        <style>
        .image-overlay {{
            position: fixed;
            top: 0;
            left: 0;
            width: 100vw;
            height: 100vh;
            background: rgba(0, 0, 0, 0.88);
            z-index: 999999;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 40px;
            box-sizing: border-box;
        }}

        .image-overlay-box {{
            max-width: 95vw;
            max-height: 92vh;
            background: #020617;
            border: 1px solid #334155;
            border-radius: 18px;
            padding: 18px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.7);
        }}

        .image-overlay-title {{
            color: #f9fafb;
            font-size: 30px;
            font-weight: 900;
            margin-bottom: 12px;
            text-align: center;
        }}

        .image-overlay-box img {{
            max-width: 90vw;
            max-height: 80vh;
            object-fit: contain;
            display: block;
            margin: auto;
            border-radius: 12px;
        }}
        </style>

        <div class="image-overlay">
            <div class="image-overlay-box">
                <div class="image-overlay-title">{naslov}</div>
                <img src="data:image/png;base64,{img64}">
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


def boja_za_masinu(masina):
    paleta = [
        "rgba(59, 130, 246, 0.18)",
        "rgba(34, 197, 94, 0.18)",
        "rgba(234, 179, 8, 0.18)",
        "rgba(168, 85, 247, 0.18)",
        "rgba(14, 165, 233, 0.18)",
        "rgba(249, 115, 22, 0.18)",
        "rgba(20, 184, 166, 0.18)",
        "rgba(236, 72, 153, 0.18)",
    ]

    if masina is None:
        return paleta[0]

    indeks = abs(hash(str(masina))) % len(paleta)
    return paleta[indeks]


def prikazi_naslov_masine(masina, iskljucena=False):
    boja = boja_za_masinu(masina)

    if iskljucena:
        border = "rgba(251, 191, 36, 0.8)"
        ikonica = "🟠"
        napomena = "<div style='font-size:12px;color:#fbbf24;margin-top:4px;'>Nije uključeno u ukupan proračun proizvodnje</div>"
    else:
        border = "rgba(148, 163, 184, 0.45)"
        ikonica = "🏭"
        napomena = ""

    st.markdown(
        f"""
        <div style="
            background: {boja};
            border: 1px solid {border};
            border-radius: 14px;
            padding: 10px 12px;
            margin-bottom: 10px;
        ">
            <div style="
                font-size: 21px;
                font-weight: 800;
                color: #ffffff;
                line-height: 1.2;
            ">
                {ikonica} {masina}
            </div>
            {napomena}
        </div>
        """,
        unsafe_allow_html=True
    )


def prikazi_grafikon(df_graf, x, y, tip_grafikona):
    if df_graf is None or df_graf.empty:
        st.info("Nema podataka za grafikon.")
        return

    if tip_grafikona == "Linijski":
        st.line_chart(df_graf, x=x, y=y)
    elif tip_grafikona == "Površinski":
        st.area_chart(df_graf, x=x, y=y)
    else:
        st.bar_chart(df_graf, x=x, y=y)


def prikazi_stubice_sa_vrednostima(df_graf, x, y, naslov=""):
    if df_graf is None or df_graf.empty:
        st.info("Nema podataka za grafikon.")
        return

    df_graf = df_graf.copy()
    df_graf[y] = pd.to_numeric(df_graf[y], errors="coerce").fillna(0)

    tekst = df_graf[y].apply(lambda v: f"{v:,.0f}".replace(",", "."))

    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=df_graf[x],
            y=df_graf[y],
            text=tekst,
            textposition="outside",
            marker=dict(line=dict(width=1)),
        )
    )

    fig.update_layout(
        title=dict(text=naslov, x=0.01, xanchor="left", font=dict(size=22, color="white")),
        height=430,
        margin=dict(l=20, r=20, t=60, b=90),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#f9fafb"),
        xaxis=dict(tickangle=-25),
        yaxis=dict(gridcolor="rgba(148,163,184,0.25)")
    )

    st.plotly_chart(fig, use_container_width=True)


def prikazi_scrap_grafikon(df_scrap, naslov="SCRAP"):
    if df_scrap is None or df_scrap.empty:
        st.info("Nema podataka za SCRAP grafikon.")
        return

    df_scrap = df_scrap.copy()

    df_scrap["NOK"] = pd.to_numeric(df_scrap["NOK"], errors="coerce").fillna(0)
    df_scrap["OK"] = pd.to_numeric(df_scrap["OK"], errors="coerce").fillna(0)

    df_scrap["Scrap_pct"] = df_scrap.apply(
        lambda r: procenat(r["NOK"], r["OK"] + r["NOK"]),
        axis=1
    )

    df_scrap["Tekst_stub"] = df_scrap["NOK"].apply(lambda x: f"{x:,.0f}".replace(",", "."))
    df_scrap["Tekst_proc"] = df_scrap["Scrap_pct"].apply(format_proc)

    fig = go.Figure()

    fig.add_trace(
        go.Bar(
            x=df_scrap["Proces"],
            y=df_scrap["NOK"],
            text=df_scrap["Tekst_stub"],
            textposition="inside",
            marker=dict(line=dict(width=1)),
            hovertemplate=(
                "<b>%{x}</b><br>"
                "NOK: %{y:,.0f}<br>"
                "Scrap: %{customdata}<extra></extra>"
            ),
            customdata=df_scrap["Tekst_proc"]
        )
    )

    for _, red in df_scrap.iterrows():
        fig.add_annotation(
            x=red["Proces"],
            y=red["NOK"],
            text=red["Tekst_proc"],
            showarrow=False,
            yshift=18,
            font=dict(size=16, color="white"),
            bgcolor="rgba(0,0,0,0.35)",
            bordercolor="rgba(255,255,255,0.45)",
            borderwidth=1
        )

    fig.update_layout(
        title=dict(text=naslov, x=0.01, xanchor="left", font=dict(size=28, color="white")),
        height=430,
        margin=dict(l=20, r=20, t=70, b=70),
        plot_bgcolor="rgba(64,64,64,0.95)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="white"),
        xaxis=dict(title="", tickfont=dict(size=14, color="white")),
        yaxis=dict(
            title="NOK komada",
            gridcolor="rgba(255,255,255,0.18)",
            zerolinecolor="rgba(255,255,255,0.35)"
        )
    )

    st.plotly_chart(fig, use_container_width=True)


def vrednost_merged_celije(ws, red, kolona):
    cell = ws.cell(row=red, column=kolona)

    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            start_cell = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            )
            return start_cell.value

    return None


def normalizuj_smenu(vrednost):
    tekst = ocisti_tekst(vrednost)

    if tekst in dozvoljene_smene:
        return tekst

    return None


def smena_za_kolonu(ws, kolona):
    vrednost = vrednost_merged_celije(ws, red_smena, kolona)
    return normalizuj_smenu(vrednost)


def poslednja_kolona_koja_pripada_smeni(ws):
    poslednja = None

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= red_smena <= merged_range.max_row:
            vrednost = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            ).value

            smena = normalizuj_smenu(vrednost)

            if smena is not None:
                if poslednja is None or merged_range.max_col > poslednja:
                    poslednja = merged_range.max_col

    for kolona in range(1, 121):
        vrednost = ws.cell(row=red_smena, column=kolona).value
        smena = normalizuj_smenu(vrednost)

        if smena is not None:
            if poslednja is None or kolona > poslednja:
                poslednja = kolona

    return poslednja


def napravi_mapu_kolona(ws):
    mapa = {}

    for kolona in range(1, 121):
        naziv = ws.cell(row=red_nazivi, column=kolona).value
        naziv_cist = ocisti_tekst(naziv)

        if naziv_cist is None:
            continue

        mapa[norm(naziv_cist)] = kolona

    return mapa


def nadji_kolonu(mapa_kolona, moguci_nazivi):
    for naziv in moguci_nazivi:
        kljuc = norm(naziv)

        if kljuc in mapa_kolona:
            return mapa_kolona[kljuc]

    return None


def suma_kolona_po_nazivu(ws, red, nazivi_kolona):
    trazeni = {norm(n) for n in nazivi_kolona}
    zadnja_kolona = poslednja_kolona_koja_pripada_smeni(ws)

    if zadnja_kolona is None:
        zadnja_kolona = 120

    ukupno = 0

    for kolona in range(1, zadnja_kolona + 1):
        naziv = ws.cell(row=red_nazivi, column=kolona).value

        if norm(naziv) in trazeni:
            ukupno += broj(ws.cell(row=red, column=kolona).value)

    return ukupno


def suma_kolona_po_kljucnim_recima(ws, red, obavezne_reci):
    zadnja_kolona = poslednja_kolona_koja_pripada_smeni(ws)

    if zadnja_kolona is None:
        zadnja_kolona = 120

    ukupno = 0
    obavezne_reci = [r.lower() for r in obavezne_reci]

    for kolona in range(1, zadnja_kolona + 1):
        naziv = norm(ws.cell(row=red_nazivi, column=kolona).value)

        if all(rec in naziv for rec in obavezne_reci):
            ukupno += broj(ws.cell(row=red, column=kolona).value)

    return ukupno


# ============================================================
# FUNKCIJE ZA NOTES
# ============================================================

def proizvod_iz_naziva(naziv):
    t = norm(naziv)

    if "stator" in t:
        return "STATOR"

    if "rotor ab" in t or "ab rotor" in t:
        return "ROTOR AB"

    if "rotor ac" in t or "ac rotor" in t:
        return "ROTOR AC"

    if "rotor ad" in t or "ad rotor" in t:
        return "ROTOR AD"

    if "rotor ae" in t or "ae rotor" in t:
        return "ROTOR AE"

    if "rotor f" in t:
        return "ROTOR F"

    if "rotor g" in t:
        return "ROTOR G"

    if "rotor" in t:
        return "ROTOR"

    return "NEPOZNATO"


def da_li_je_nok_kolona(naziv):
    t = norm(naziv)
    return "nok" in t


def da_li_je_stops_kolona(naziv):
    t = norm(naziv)
    return "stops" in t and "min" in t


def podeli_note_na_stavke(note):
    if note is None:
        return []

    tekst = str(note).replace("\r", "\n")
    delovi = []

    for linija in tekst.split("\n"):
        linija = linija.strip()
        if da_li_je_excel_threaded_comment_poruka(linija):
            continue
        if linija:
            delovi.append(linija)

    if not delovi and tekst.strip():
        delovi = [tekst.strip()]

    return delovi


def parsiraj_nok_note(note):
    stavke = podeli_note_na_stavke(note)
    rezultat = []

    pattern = re.compile(
        r"^\s*(\d+(?:[.,]\d+)?)\s*(?:kom|komada|pcs|pc|x|×)?\s*(.*)$",
        re.IGNORECASE
    )

    for stavka in stavke:
        if da_li_je_excel_threaded_comment_poruka(stavka) or da_li_je_samo_ime_kolege(stavka):
            continue

        m = pattern.match(stavka)

        if m:
            kolicina_txt = m.group(1).replace(",", ".")
            razlog_original = m.group(2).strip(" -–—:;,.")

            try:
                kolicina = float(kolicina_txt)
            except Exception:
                kolicina = None

            if razlog_original == "":
                razlog_original = "nije upisan razlog"
        else:
            kolicina = None
            razlog_original = stavka.strip(" -–—:;,.")

        rezultat.append({
            "Komada_iz_note": kolicina,
            "Razlog": normalizuj_razlog(razlog_original),
            "Originalna_stavka": stavka
        })

    return rezultat


def da_li_je_samo_ime_kolege(tekst):
    """
    Vraća True ako je cela stavka iz note-a samo ime kolege/autora.
    Takve redove ne treba pretvarati u "nije upisan razlog", jer tada
    program pogrešno dodeli celu STOP/MIN vrednost imenu kolege.
    """
    original = str(tekst or "")
    ocisceno = finalno_ocisti_razlog_od_imena(original)
    ocisceno = re.sub(r"^[\s\.\,\;\:\-\–\—\_`]+", "", ocisceno).strip()
    ocisceno = re.sub(r"[\s\.\,\;\:\-\–\—\_`]+$", "", ocisceno).strip()

    ima_ime = any(
        re.search(r"(?<!\w)" + re.escape(rec) + r"(?!\w)", original, flags=re.IGNORECASE)
        for rec in RECI_KOJE_SE_BRISU_IZ_RAZLOGA
    )

    return ima_ime and ocisceno == ""


def parsiraj_stop_note(note, vrednost_celije):
    stavke = podeli_note_na_stavke(note)
    rezultat = []

    # 1) Standardni unos: 30min razlog, 30 min razlog, 30m razlog, 30' razlog.
    # Dodato je i "mi" zbog grešaka u unosu kao "55mi otisak".
    pattern_minuti = re.compile(
        r"(?<!\d)(\d+(?:[.,]\d+)?)\s*(?:minuta|minut|min|minute|mi\b|m\b|'|′)",
        re.IGNORECASE
    )

    # 2) Unos sa apostrofom ispred broja: '30 razlog.
    pattern_apostrof_pre = re.compile(
        r"(?:'|′)\s*(\d+(?:[.,]\d+)?)",
        re.IGNORECASE
    )

    # 3) Čest unos bez jedinice: 40 messfeed, 15 otisak, 10 zaglavljen lim.
    # Ovo se koristi samo ako linija počinje brojem i tekstom.
    pattern_broj_bez_jedinice = re.compile(
        r"^\s*[`'′\-–—]*\s*(\d+(?:[.,]\d+)?)\s+(.+\S)\s*$",
        re.IGNORECASE
    )

    for stavka in stavke:
        stavka = str(stavka).strip()

        if stavka == "":
            continue

        if da_li_je_excel_threaded_comment_poruka(stavka):
            continue

        # Autor note-a, npr. "Bojan Smiljanić" ili "Dimitrije Miler", nije razlog zastoja.
        if da_li_je_samo_ime_kolege(stavka):
            continue

        m = pattern_minuti.search(stavka)
        koristi_apostrof_pre = False

        if not m:
            m = pattern_apostrof_pre.search(stavka)
            koristi_apostrof_pre = m is not None

        if m:
            minuti_txt = m.group(1).replace(",", ".")

            try:
                minuti = float(minuti_txt)
            except Exception:
                minuti = None

            if koristi_apostrof_pre:
                razlog_original = pattern_apostrof_pre.sub("", stavka, count=1).strip(" -–—:;,.`")
            else:
                razlog_original = pattern_minuti.sub("", stavka, count=1).strip(" -–—:;,.`")

            razlog = normalizuj_razlog(razlog_original)

            if razlog != "nije upisan razlog":
                rezultat.append({
                    "Minuta_iz_note": minuti,
                    "Razlog": razlog,
                    "Originalna_stavka": stavka
                })

            continue

        m = pattern_broj_bez_jedinice.match(stavka)

        if m and not re.search(r"\d{1,2}:\d{2}|\d{1,2}\.\d{2}", stavka):
            minuti_txt = m.group(1).replace(",", ".")

            try:
                minuti = float(minuti_txt)
            except Exception:
                minuti = None

            razlog = normalizuj_razlog(m.group(2))

            if razlog != "nije upisan razlog":
                rezultat.append({
                    "Minuta_iz_note": minuti,
                    "Razlog": razlog,
                    "Originalna_stavka": stavka
                })

            continue

        # Ako nema upisanih minuta u liniji, ali postoji konkretan tekst razloga,
        # koristi se vrednost iz STOP/MIN ćelije. Ovo pokriva unos tipa "otisak".
        razlog = normalizuj_razlog(stavka)

        if razlog == "nije upisan razlog":
            continue

        rezultat.append({
            "Minuta_iz_note": broj(vrednost_celije),
            "Razlog": razlog,
            "Originalna_stavka": stavka
        })

    return rezultat


# ============================================================
# KPI FUNKCIJE
# ============================================================

def izracunaj_kpi(df_kpi):
    if df_kpi is None or df_kpi.empty:
        return {
            "Opening_time_min": 0,
            "Stops_min": 0,
            "Availability_pct": None,
            "Plan": 0,
            "Realizacija": 0,
            "Realizacija_pct": None,
            "OK": 0,
            "NOK": 0,
            "Scrap_rate_pct": None,
            "Quality_pct": None,
            "OEE_pct": None,
        }

    df_kpi = df_kpi.copy()

    for kol in [
        "Opening_time_min",
        "Stops_min",
        "Plan_STATOR",
        "Plan_ROTOR",
        "Realizacija_STATOR",
        "Realizacija_ROTOR",
        "OK_STATOR",
        "OK_ROTOR",
        "NOK_STATOR",
        "NOK_ROTOR",
    ]:
        if kol not in df_kpi.columns:
            df_kpi[kol] = 0

        df_kpi[kol] = pd.to_numeric(df_kpi[kol], errors="coerce").fillna(0)

    opening = df_kpi["Opening_time_min"].sum()
    stops = df_kpi["Stops_min"].sum()

    plan = df_kpi["Plan_STATOR"].sum() + df_kpi["Plan_ROTOR"].sum()
    realizacija = df_kpi["Realizacija_STATOR"].sum() + df_kpi["Realizacija_ROTOR"].sum()

    ok = df_kpi["OK_STATOR"].sum() + df_kpi["OK_ROTOR"].sum()
    nok = df_kpi["NOK_STATOR"].sum() + df_kpi["NOK_ROTOR"].sum()

    availability = procenat(opening - stops, opening)
    realizacija_pct = procenat(realizacija, plan)
    scrap_rate = procenat(nok, ok + nok)

    quality = None
    if scrap_rate is not None:
        quality = 100 - scrap_rate

    oee = None
    if availability is not None and realizacija_pct is not None and quality is not None:
        oee = availability * realizacija_pct * quality / 10000

    return {
        "Opening_time_min": opening,
        "Stops_min": stops,
        "Availability_pct": availability,
        "Plan": plan,
        "Realizacija": realizacija,
        "Realizacija_pct": realizacija_pct,
        "OK": ok,
        "NOK": nok,
        "Scrap_rate_pct": scrap_rate,
        "Quality_pct": quality,
        "OEE_pct": oee,
    }


def prikazi_kpi_karticu(naslov, df_kpi, iskljucena=False):
    kpi = izracunaj_kpi(df_kpi)

    with st.container(border=True):
        prikazi_naslov_masine(naslov, iskljucena=iskljucena)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Availability", format_proc(kpi["Availability_pct"]))
        c2.metric("Realizacija", format_proc(kpi["Realizacija_pct"]))
        c3.metric("Scrap rate", format_proc(kpi["Scrap_rate_pct"]))
        c4.metric("OEE", format_proc(kpi["OEE_pct"]))

        c5, c6, c7, c8 = st.columns(4)
        c5.metric("Opening time/min", format_broj(kpi["Opening_time_min"]))
        c6.metric("Stops/min", format_broj(kpi["Stops_min"]))
        c7.metric("Plan", format_broj(kpi["Plan"]))
        c8.metric("Realizacija kom", format_broj(kpi["Realizacija"]))

        c9, c10, c11, c12 = st.columns(4)
        c9.metric("OK", format_broj(kpi["OK"]))
        c10.metric("NOK", format_broj(kpi["NOK"]))
        c11.metric("Quality", format_proc(kpi["Quality_pct"]))
        c12.metric("Proces/Projekat", naslov)

        availability = 0 if kpi["Availability_pct"] is None else float(kpi["Availability_pct"])
        realizacija = 0 if kpi["Realizacija_pct"] is None else float(kpi["Realizacija_pct"])
        quality = 0 if kpi["Quality_pct"] is None else float(kpi["Quality_pct"])
        oee = 0 if kpi["OEE_pct"] is None else float(kpi["OEE_pct"])

        nazivi = ["Availability", "Realizacija", "Quality", "OEE"]
        vrednosti = [availability, realizacija, quality, oee]

        tip = "Gauge"

        if tip == "Radar":
            fig = go.Figure(
                go.Scatterpolar(
                    r=vrednosti + [vrednosti[0]],
                    theta=nazivi + [nazivi[0]],
                    fill="toself",
                    line=dict(color="#22D3EE", width=3),
                    fillcolor="rgba(34,211,238,0.20)",
                    marker=dict(size=8, color="#A78BFA"),
                    hovertemplate="%{theta}: %{r:.1f}%<extra></extra>"
                )
            )

            fig.update_layout(
                polar=dict(
                    bgcolor="rgba(15,23,42,0.55)",
                    radialaxis=dict(
                        visible=True,
                        range=[0, max(110, max(vrednosti) * 1.15 if vrednosti else 110)],
                        ticksuffix="%",
                        gridcolor="rgba(148,163,184,0.22)",
                        linecolor="rgba(148,163,184,0.25)"
                    ),
                    angularaxis=dict(
                        gridcolor="rgba(148,163,184,0.22)",
                        linecolor="rgba(148,163,184,0.25)"
                    )
                )
            )

        elif tip == "Gauge":
            fig = go.Figure()

            gauge_podaci = [
                ("Availability", availability, [0.08, 0.45], [0.55, 0.92]),
                ("Realizacija", realizacija, [0.55, 0.92], [0.55, 0.92]),
                ("Quality", quality, [0.08, 0.45], [0.05, 0.42]),
                ("OEE", oee, [0.55, 0.92], [0.05, 0.42]),
            ]

            for naziv_g, vrednost_g, x_domen, y_domen in gauge_podaci:
                fig.add_trace(
                    go.Indicator(
                        mode="gauge+number",
                        value=vrednost_g,
                        number=dict(suffix="%", font=dict(size=28, color="#F8FAFC")),
                        title=dict(text=naziv_g, font=dict(size=16, color="#CBD5E1")),
                        domain=dict(x=x_domen, y=y_domen),
                        gauge=dict(
                            axis=dict(range=[0, max(110, vrednost_g * 1.15)], tickcolor="#94A3B8"),
                            bar=dict(color="#22D3EE"),
                            bgcolor="rgba(15,23,42,0.55)",
                            borderwidth=0,
                            steps=[
                                dict(range=[0, 70], color="rgba(239,68,68,0.15)"),
                                dict(range=[70, 90], color="rgba(249,115,22,0.15)"),
                                dict(range=[90, 100], color="rgba(59,130,246,0.15)"),
                                dict(range=[100, max(110, vrednost_g * 1.15)], color="rgba(34,197,94,0.15)")
                            ],
                            threshold=dict(
                                line=dict(color="#F8FAFC", width=3),
                                thickness=0.75,
                                value=100
                            )
                        )
                    )
                )

        else:
            boje = [
                "#38BDF8",
                "#A78BFA",
                "#22C55E",
                "#F59E0B"
            ]

            fig = go.Figure(
                go.Bar(
                    x=nazivi,
                    y=vrednosti,
                    text=[f"{v:.1f}%" for v in vrednosti],
                    textposition="outside",
                    cliponaxis=False,
                    marker=dict(
                        color=boje,
                        line=dict(color="rgba(255,255,255,0.30)", width=1),
                        cornerradius=10
                    ),
                    hovertemplate="<b>%{x}</b><br>%{y:.1f}%<extra></extra>"
                )
            )

            fig.add_hline(
                y=100,
                line_dash="dash",
                line_color="#22C55E",
                annotation_text="Cilj 100%",
                annotation_position="top left",
                annotation_font_color="#86EFAC"
            )

            fig.update_yaxes(
                range=[0, max(115, max(vrednosti) * 1.20 if vrednosti else 115)],
                ticksuffix="%",
                gridcolor="rgba(148,163,184,0.18)"
            )

        fig.update_layout(
            title=dict(
                text=f"KPI globalni — {naslov}",
                x=0.01,
                xanchor="left",
                font=dict(size=22, color="#F8FAFC")
            ),
            height=420 if tip != "Gauge" else 520,
            margin=dict(l=25, r=25, t=70, b=45),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(15,23,42,0.40)",
            font=dict(color="#E2E8F0"),
            showlegend=False,
            hoverlabel=dict(
                bgcolor="#0F172A",
                font_color="#F8FAFC",
                bordercolor="#475569"
            )
        )

        st.plotly_chart(
            fig,
            use_container_width=True,
            config={"displayModeBar": False}
        )
# ============================================================
# GLAVNO UČITAVANJE
# ============================================================

@st.cache_data(show_spinner="Učitavam podatke iz izabranog Excel fajla...")
def ucitaj_sve_podatke(original_fajl):
    if not os.path.exists(original_fajl):
        raise FileNotFoundError(
            f"Fajl nije pronađen. Proveri VPN, R: disk i naziv fajla:\n{original_fajl}"
        )

    folder_kopije = tempfile.gettempdir()
    os.makedirs(folder_kopije, exist_ok=True)

    vreme = datetime.now().strftime("%Y%m%d_%H%M%S")
    kopija_fajla = os.path.join(folder_kopije, f"proizvodnja_kopija_dashboard_{vreme}.xlsx")

    shutil.copy2(original_fajl, kopija_fajla)

    wb = load_workbook(kopija_fajla, data_only=True)

    dnevni_redovi = []
    nok_redovi = []
    zastoji_redovi = []

    for ws in wb.worksheets:
        naziv_taba = ws.title.strip()

        if naziv_taba in preskoci_tabove:
            continue

        mapa_kolona = napravi_mapu_kolona(ws)

        kolone = {}
        for naziv_izlaza, moguci_nazivi in MAPIRANJE.items():
            kolone[naziv_izlaza] = nadji_kolonu(mapa_kolona, moguci_nazivi)

        for red in range(prvi_red_podataka, poslednji_red_podataka + 1):
            datum = ws.cell(row=red, column=kolona_datum).value

            if datum is None:
                continue

            zapis = {
                "Datum": datum,
                "Masina": naziv_taba,
            }

            for naziv_izlaza, kolona in kolone.items():
                if kolona is None:
                    zapis[naziv_izlaza] = None
                    continue

                cell = ws.cell(row=red, column=kolona)

                if naziv_izlaza == "Reason":
                    vrednost = cell.value

                    if cell.comment:
                        note = ocisti_note(cell.comment.text)
                        if note:
                            if vrednost:
                                vrednost = f"{vrednost} | NOTE: {note}"
                            else:
                                vrednost = note

                    zapis[naziv_izlaza] = vrednost
                else:
                    zapis[naziv_izlaza] = broj(cell.value)

            zapis["Opening_time_min"] = suma_kolona_po_kljucnim_recima(
                ws,
                red,
                ["opening", "time"]
            )

            zapis["Stops_min"] = suma_kolona_po_kljucnim_recima(
                ws,
                red,
                ["stops", "min"]
            )

            zapis["Work_time_min"] = zapis["Opening_time_min"] - zapis["Stops_min"]

            # Posebno privremeno pravilo za AIDA VITESKO:
            # dok tabela nije sređena, koristi se samo kolona G kao realizacija,
            # a notes analiza se i dalje radi iz smenskih kolona.
            if naziv_taba.upper() == "AIDA VITESKO":
                zapis["Plan_STATOR"] = 0
                zapis["Plan_ROTOR"] = 0
                zapis["Realizacija_STATOR"] = broj(ws.cell(row=red, column=7).value)
                zapis["Realizacija_ROTOR"] = 0
                zapis["OK_STATOR"] = 0
                zapis["OK_ROTOR"] = 0
                zapis["NOK_STATOR"] = 0
                zapis["NOK_ROTOR"] = 0

            # Posebno pravilo za HEATING VITESCO tabove:
            # u redu 7 piše TOTAL OK PARTS bez STATOR/ROTOR u nazivu kolone.
            if "HEATING VITESCO STATOR" in naziv_taba.upper():
                kol_total_ok = nadji_kolonu(
                    mapa_kolona,
                    [
                        "TOTAL OK PARTS",
                        "Total OK Parts",
                        "Total OK parts",
                    ]
                )

                kol_total_scrap = nadji_kolonu(
                    mapa_kolona,
                    [
                        "TOTAL SCRAP",
                        "Total scrap",
                        "Total Scrap",
                    ]
                )

                if kol_total_ok is not None:
                    zapis["OK_STATOR"] = broj(ws.cell(row=red, column=kol_total_ok).value)

                if kol_total_scrap is not None:
                    zapis["NOK_STATOR"] = broj(ws.cell(row=red, column=kol_total_scrap).value)

            if "HEATING VITESCO ROTOR" in naziv_taba.upper():
                kol_total_ok = nadji_kolonu(
                    mapa_kolona,
                    [
                        "TOTAL OK PARTS",
                        "Total OK Parts",
                        "Total OK parts",
                    ]
                )

                kol_total_scrap = nadji_kolonu(
                    mapa_kolona,
                    [
                        "TOTAL SCRAP",
                        "Total scrap",
                        "Total Scrap",
                    ]
                )

                if kol_total_ok is not None:
                    zapis["OK_ROTOR"] = broj(ws.cell(row=red, column=kol_total_ok).value)

                if kol_total_scrap is not None:
                    zapis["NOK_ROTOR"] = broj(ws.cell(row=red, column=kol_total_scrap).value)

            # Posebno pravilo za tabove gde se ROTOR OK/NOK čita iz smenskih kolona.
            # Za sada je to DMC EMR4 ROTOR.
            if naziv_taba in rotor_iz_smena_tabovi:
                ok_rotor_smena = suma_kolona_po_nazivu(
                    ws,
                    red,
                    [
                        "Good Parts/rotor",
                        "Good Parts /rotor",
                        "Good Parts / rotor",
                        "Good Parts rotor",
                    ]
                )

                nok_rotor_smena = suma_kolona_po_nazivu(
                    ws,
                    red,
                    [
                        "NOK parts rotor",
                        "NOK parts/rotor",
                        "NOK parts /rotor",
                        "NOK parts / rotor",
                    ]
                )

                zapis["OK_ROTOR"] = ok_rotor_smena
                zapis["NOK_ROTOR"] = nok_rotor_smena

                if broj(zapis.get("Realizacija_ROTOR")) == 0:
                    zapis["Realizacija_ROTOR"] = ok_rotor_smena + nok_rotor_smena

            if naziv_taba in dmc_stator_iz_smena_tabovi:
                ok_dmc = suma_kolona_po_nazivu(
                    ws,
                    red,
                    [
                        "Good Parts/stator",
                        "Good Parts /stator",
                        "Good Parts / stator",
                        "Good Parts stator",
                    ]
                )

                nok_dmc = suma_kolona_po_nazivu(
                    ws,
                    red,
                    [
                        "NOK parts stator",
                        "NOK parts/stator",
                        "NOK parts /stator",
                        "NOK parts / stator",
                    ]
                )

                zapis["OK_STATOR"] = ok_dmc
                zapis["NOK_STATOR"] = nok_dmc

                if broj(zapis.get("Realizacija_STATOR")) == 0:
                    zapis["Realizacija_STATOR"] = ok_dmc + nok_dmc


            # === START EMR4 POSEBNO PRAVILO ===
            # Posebno pravilo za DMC EMR4 tabove:
            # U tim tabovima nazivi kolona nisu dovoljno jasni za automatsko mapiranje.
            #
            # DMC EMR4 STATOR:
            # C = Plan stator
            # E = TOTAL Realization stator
            # F = TOTAL OK PARTS
            # W = TOTAL SCRAP
            #
            # DMC EMR4 ROTOR:
            # C = Plan rotor, iako u fajlu može da piše Plan stator
            # E = TOTAL Realization rotor
            # F = TOTAL OK PARTS
            # W = TOTAL SCRAP
            #
            # Ako W nema vrednost, NOK se sabira iz smenskih NOK kolona:
            # L, P, T

            if naziv_taba.upper() == "DMC EMR4 STATOR":
                nok_stator_total = broj(ws.cell(row=red, column=23).value)  # W = TOTAL SCRAP

                if nok_stator_total == 0:
                    nok_stator_total = (
                        broj(ws.cell(row=red, column=12).value) +  # L = I shift NOK
                        broj(ws.cell(row=red, column=16).value) +  # P = II shift NOK
                        broj(ws.cell(row=red, column=20).value)    # T = III shift NOK
                    )

                zapis["Plan_STATOR"] = broj(ws.cell(row=red, column=3).value)          # C
                zapis["Realizacija_STATOR"] = broj(ws.cell(row=red, column=5).value)   # E
                zapis["OK_STATOR"] = broj(ws.cell(row=red, column=6).value)            # F
                zapis["NOK_STATOR"] = nok_stator_total

                zapis["Plan_ROTOR"] = 0
                zapis["Realizacija_ROTOR"] = 0
                zapis["OK_ROTOR"] = 0
                zapis["NOK_ROTOR"] = 0

            if naziv_taba.upper() == "DMC EMR4 ROTOR":
                nok_rotor_total = broj(ws.cell(row=red, column=23).value)  # W = TOTAL SCRAP

                if nok_rotor_total == 0:
                    nok_rotor_total = (
                        broj(ws.cell(row=red, column=12).value) +  # L = I shift NOK
                        broj(ws.cell(row=red, column=16).value) +  # P = II shift NOK
                        broj(ws.cell(row=red, column=20).value)    # T = III shift NOK
                    )

                zapis["Plan_STATOR"] = 0
                zapis["Realizacija_STATOR"] = 0
                zapis["OK_STATOR"] = 0
                zapis["NOK_STATOR"] = 0

                zapis["Plan_ROTOR"] = broj(ws.cell(row=red, column=3).value)           # C
                zapis["Realizacija_ROTOR"] = broj(ws.cell(row=red, column=5).value)    # E
                zapis["OK_ROTOR"] = broj(ws.cell(row=red, column=6).value)             # F
                zapis["NOK_ROTOR"] = nok_rotor_total
            # === END EMR4 POSEBNO PRAVILO ===

            rotor_po_tipovima = (
                broj(zapis.get("Total_production_AB")) +
                broj(zapis.get("Total_production_AC")) +
                broj(zapis.get("Total_production_AD")) +
                broj(zapis.get("Total_production_AE"))
            )

            if broj(zapis.get("Realizacija_ROTOR")) == 0 and rotor_po_tipovima > 0:
                zapis["Realizacija_ROTOR"] = rotor_po_tipovima

            ok_stator = broj(zapis.get("OK_STATOR"))
            nok_stator = broj(zapis.get("NOK_STATOR"))
            ok_rotor = broj(zapis.get("OK_ROTOR"))
            nok_rotor = broj(zapis.get("NOK_ROTOR"))

            plan_stator = broj(zapis.get("Plan_STATOR"))
            plan_rotor = broj(zapis.get("Plan_ROTOR"))

            realizacija_stator = broj(zapis.get("Realizacija_STATOR"))
            realizacija_rotor = broj(zapis.get("Realizacija_ROTOR"))

            zapis["Ukupno_STATOR"] = ok_stator + nok_stator
            zapis["Ukupno_ROTOR"] = ok_rotor + nok_rotor

            zapis["NOK_proc_STATOR"] = procenat(nok_stator, ok_stator + nok_stator)
            zapis["NOK_proc_ROTOR"] = procenat(nok_rotor, ok_rotor + nok_rotor)

            zapis["Realizacija_proc_STATOR"] = procenat(realizacija_stator, plan_stator)
            zapis["Realizacija_proc_ROTOR"] = procenat(realizacija_rotor, plan_rotor)

            dnevni_redovi.append(zapis)

        # NOTES ANALIZA
        zadnja_kolona = poslednja_kolona_koja_pripada_smeni(ws)

        if zadnja_kolona is not None:
            for kolona in range(1, zadnja_kolona + 1):
                naziv_podatka = ocisti_tekst(ws.cell(row=red_nazivi, column=kolona).value)

                if naziv_podatka is None:
                    continue

                smena = smena_za_kolonu(ws, kolona)
                proizvod = proizvod_iz_naziva(naziv_podatka)
                tip_proizvoda = tip_proizvoda_iz_proizvoda(proizvod)

                jeste_nok = da_li_je_nok_kolona(naziv_podatka)
                jeste_stop = da_li_je_stops_kolona(naziv_podatka)

                # Sigurnosno pravilo za DMC EMR4 STATOR:
                # kolone L, P i T su NOK stator kolone sa notesima.
                if naziv_taba.upper() == "DMC EMR4 STATOR" and kolona in [12, 16, 20]:
                    jeste_nok = True
                    proizvod = "STATOR"
                    tip_proizvoda = "STATOR"

                # Sigurnosno pravilo za DMC EMR4 ROTOR:
                # kolone L, P i T su NOK rotor kolone sa notesima.
                if naziv_taba.upper() == "DMC EMR4 ROTOR" and kolona in [12, 16, 20]:
                    jeste_nok = True
                    proizvod = "ROTOR"
                    tip_proizvoda = "ROTOR"

                if not jeste_nok and not jeste_stop:
                    continue

                for red in range(prvi_red_podataka, poslednji_red_podataka + 1):
                    datum = ws.cell(row=red, column=kolona_datum).value

                    if datum is None:
                        continue

                    cell = ws.cell(row=red, column=kolona)
                    note = ocisti_note(cell.comment.text) if cell.comment else None

                    if note is None or str(note).strip() == "":
                        continue

                    vrednost_celije = cell.value

                    if jeste_nok:
                        stavke = parsiraj_nok_note(note)

                        for stavka in stavke:
                            nok_redovi.append({
                                "Datum": datum,
                                "Masina": naziv_taba,
                                "Smena": smena,
                                "Kolona": cell.coordinate,
                                "Naziv_podatka": naziv_podatka,
                                "Proizvod": proizvod,
                                "Tip_proizvoda": tip_proizvoda,
                                "Vrednost_celije_NOK": vrednost_celije,
                                "Komada_iz_note": stavka["Komada_iz_note"],
                                "Razlog": stavka["Razlog"],
                                "Originalna_stavka": stavka["Originalna_stavka"],
                                "Originalni_note": note
                            })

                    if jeste_stop:
                        stavke = parsiraj_stop_note(note, vrednost_celije)

                        for stavka in stavke:
                            zastoji_redovi.append({
                                "Datum": datum,
                                "Masina": naziv_taba,
                                "Smena": smena,
                                "Kolona": cell.coordinate,
                                "Naziv_podatka": naziv_podatka,
                                "Vrednost_celije_stops_min": vrednost_celije,
                                "Minuta_iz_note": stavka["Minuta_iz_note"],
                                "Razlog": stavka["Razlog"],
                                "Originalna_stavka": stavka["Originalna_stavka"],
                                "Originalni_note": note
                            })

    df_dnevni = pd.DataFrame(dnevni_redovi)
    df_nok = pd.DataFrame(nok_redovi)
    df_nok = finalno_ocisti_df_razloge(df_nok)
    df_zastoji = pd.DataFrame(zastoji_redovi)
    df_zastoji = finalno_ocisti_df_razloge(df_zastoji)

    if not df_dnevni.empty:
        df_dnevni["Datum"] = pd.to_datetime(df_dnevni["Datum"], errors="coerce")

    if not df_nok.empty:
        df_nok["Datum"] = pd.to_datetime(df_nok["Datum"], errors="coerce")
        df_nok["Komada_iz_note"] = pd.to_numeric(
            df_nok["Komada_iz_note"],
            errors="coerce"
        ).fillna(0)

    if not df_zastoji.empty:
        df_zastoji["Datum"] = pd.to_datetime(df_zastoji["Datum"], errors="coerce")
        df_zastoji["Minuta_iz_note"] = pd.to_numeric(
            df_zastoji["Minuta_iz_note"],
            errors="coerce"
        ).fillna(0)


    # ========================================================
    # PROCESNI NOK - ISKLJUČIVANJE IZ SVIH NOK PRORAČUNA
    # ========================================================

    if not df_nok.empty:
        procesni_obrazac = r"\bprocesn(?:i|o|ih)\b"

        maska_procesni = (
            df_nok["Originalna_stavka"].astype(str).str.contains(
                procesni_obrazac, case=False, na=False, regex=True
            )
            |
            df_nok["Razlog"].astype(str).str.contains(
                procesni_obrazac, case=False, na=False, regex=True
            )
        )

        df_procesni_nok = df_nok[maska_procesni].copy()

        if not df_procesni_nok.empty and not df_dnevni.empty:
            df_procesni_nok["Komada_iz_note"] = pd.to_numeric(
                df_procesni_nok["Komada_iz_note"], errors="coerce"
            ).fillna(0)

            korekcije = (
                df_procesni_nok
                .groupby(["Datum", "Masina", "Tip_proizvoda"], as_index=False)["Komada_iz_note"]
                .sum()
            )

            for korekcija in korekcije.itertuples(index=False):
                maska_dnevni = (
                    (df_dnevni["Datum"] == korekcija.Datum) &
                    (df_dnevni["Masina"] == korekcija.Masina)
                )

                if korekcija.Tip_proizvoda == "STATOR":
                    kolona_nok = "NOK_STATOR"
                elif korekcija.Tip_proizvoda == "ROTOR":
                    kolona_nok = "NOK_ROTOR"
                else:
                    continue

                if kolona_nok not in df_dnevni.columns:
                    continue

                trenutno = pd.to_numeric(
                    df_dnevni.loc[maska_dnevni, kolona_nok], errors="coerce"
                ).fillna(0)

                df_dnevni.loc[maska_dnevni, kolona_nok] = (
                    trenutno - korekcija.Komada_iz_note
                ).clip(lower=0)

            for kolona in ["OK_STATOR", "OK_ROTOR", "NOK_STATOR", "NOK_ROTOR"]:
                if kolona not in df_dnevni.columns:
                    df_dnevni[kolona] = 0
                df_dnevni[kolona] = pd.to_numeric(
                    df_dnevni[kolona], errors="coerce"
                ).fillna(0)

            df_dnevni["Ukupno_STATOR"] = df_dnevni["OK_STATOR"] + df_dnevni["NOK_STATOR"]
            df_dnevni["Ukupno_ROTOR"] = df_dnevni["OK_ROTOR"] + df_dnevni["NOK_ROTOR"]

            df_dnevni["NOK_proc_STATOR"] = df_dnevni.apply(
                lambda red: procenat(red["NOK_STATOR"], red["OK_STATOR"] + red["NOK_STATOR"]),
                axis=1
            )
            df_dnevni["NOK_proc_ROTOR"] = df_dnevni.apply(
                lambda red: procenat(red["NOK_ROTOR"], red["OK_ROTOR"] + red["NOK_ROTOR"]),
                axis=1
            )

        df_nok = df_nok[~maska_procesni].copy()


    return df_dnevni, df_nok, df_zastoji



# ============================================================
# ONLINE / MODERAN UI - UPLOAD FAJLA + TOP TOOLBAR + LEVI TABOVI
# ============================================================

# Sklanjamo stari naslov da dashboard izgleda kao aplikacija.
st.markdown(
    """
    <style>
    .stApp {
        background:
            radial-gradient(circle at 10% 10%, rgba(56, 189, 248, 0.12), transparent 30%),
            radial-gradient(circle at 90% 5%, rgba(168, 85, 247, 0.13), transparent 28%),
            linear-gradient(135deg, #020617 0%, #07111f 45%, #030712 100%) !important;
        color: #e5e7eb !important;
    }

    .block-container {
        padding-top: 1.0rem !important;
        padding-left: 1.4rem !important;
        padding-right: 1.4rem !important;
        max-width: 100% !important;
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, rgba(2, 6, 23, 0.98), rgba(15, 23, 42, 0.94)) !important;
        border-right: 1px solid rgba(56, 189, 248, 0.25) !important;
        box-shadow: 12px 0 35px rgba(0, 0, 0, 0.35);
    }

    section[data-testid="stSidebar"] * {
        color: #e5e7eb !important;
    }

    h1, h2, h3, h4, h5, h6, p, label, span, div {
        color: #e5e7eb;
    }

    .toolbar-card {
        background: linear-gradient(135deg, rgba(15, 23, 42, 0.94), rgba(2, 6, 23, 0.92));
        border: 1px solid rgba(56, 189, 248, 0.28);
        border-radius: 16px;
        padding: 8px 12px 2px 12px;
        margin-bottom: 10px;
        box-shadow: 0 12px 34px rgba(0, 0, 0, 0.30), inset 0 1px 0 rgba(255,255,255,0.05);
    }

    .toolbar-card label {
        font-size: 12px !important;
        font-weight: 800 !important;
        color: #bfdbfe !important;
        margin-bottom: 1px !important;
    }

    .toolbar-card [data-testid="stVerticalBlock"] {
        gap: 0.15rem !important;
    }

    .toolbar-card [data-testid="stFileUploader"] {
        margin-top: -4px !important;
    }

    .toolbar-card [data-testid="stFileUploader"] section {
        min-height: 38px !important;
        padding: 4px 8px !important;
        border-radius: 12px !important;
    }

    .hero-title {
        font-size: 34px;
        font-weight: 950;
        letter-spacing: -0.6px;
        margin-bottom: 2px;
        color: #ffffff !important;
        text-shadow: 0 0 18px rgba(56, 189, 248, 0.25);
    }

    .hero-sub {
        color: #93c5fd !important;
        margin-bottom: 14px;
        font-size: 14px;
    }

    .neo-card {
        background: linear-gradient(145deg, rgba(15, 23, 42, 0.96), rgba(2, 6, 23, 0.94));
        border: 1px solid rgba(56, 189, 248, 0.22);
        border-radius: 22px;
        padding: 18px 20px;
        margin: 10px 0;
        box-shadow: 0 18px 45px rgba(0,0,0,0.34), inset 0 1px 0 rgba(255,255,255,0.06);
    }

    .neo-card-green {
        background: linear-gradient(145deg, rgba(6, 78, 59, 0.36), rgba(2, 6, 23, 0.94));
        border-color: rgba(16, 185, 129, 0.38);
    }

    .neo-card-purple {
        background: linear-gradient(145deg, rgba(88, 28, 135, 0.36), rgba(2, 6, 23, 0.94));
        border-color: rgba(168, 85, 247, 0.38);
    }

    .neo-card-orange {
        background: linear-gradient(145deg, rgba(124, 45, 18, 0.33), rgba(2, 6, 23, 0.94));
        border-color: rgba(251, 146, 60, 0.38);
    }

    .metric-label-modern {
        color: #93c5fd !important;
        font-size: 12px;
        text-transform: uppercase;
        font-weight: 900;
        letter-spacing: 1.1px;
        margin-bottom: 6px;
    }

    .metric-value-modern {
        color: #ffffff !important;
        font-size: 30px;
        font-weight: 950;
        line-height: 1.1;
    }

    .metric-sub-modern {
        color: #cbd5e1 !important;
        font-size: 13px;
        margin-top: 6px;
    }

    .machine-title {
        font-size: 22px;
        font-weight: 950;
        color: #ffffff !important;
        margin-bottom: 4px;
    }

    .pill {
        display: inline-block;
        padding: 5px 10px;
        margin-right: 6px;
        border-radius: 999px;
        background: rgba(14, 165, 233, 0.16);
        border: 1px solid rgba(56, 189, 248, 0.36);
        color: #bfdbfe !important;
        font-weight: 800;
        font-size: 12px;
    }

    .reason-card {
        background: rgba(2, 6, 23, 0.76);
        border: 1px solid rgba(71, 85, 105, 0.75);
        border-left: 5px solid #38bdf8;
        border-radius: 16px;
        padding: 13px 16px;
        margin: 8px 0;
    }

    .reason-title {
        font-size: 17px;
        font-weight: 900;
        color: #ffffff !important;
    }

    .reason-sub {
        color: #bfdbfe !important;
        font-size: 13px;
        font-weight: 700;
        margin-top: 5px;
    }

    div[data-testid="stMetric"] {
        background: linear-gradient(145deg, rgba(15,23,42,.95), rgba(2,6,23,.92)) !important;
        border: 1px solid rgba(56,189,248,.22) !important;
        border-radius: 18px !important;
        box-shadow: 0 14px 35px rgba(0,0,0,.28);
    }

    div[data-testid="stDataFrame"] {
        background: rgba(2, 6, 23, 0.95) !important;
        border: 1px solid rgba(56, 189, 248, 0.20) !important;
        border-radius: 18px !important;
        overflow: hidden !important;
    }

    .stDataFrame div, .stDataFrame span {
        color: #e5e7eb !important;
    }

    [data-testid="stFileUploader"] section {
        background: rgba(15, 23, 42, 0.8) !important;
        border: 1px dashed rgba(56, 189, 248, 0.55) !important;
        border-radius: 18px !important;
    }

    div[data-baseweb="select"] > div,
    div[data-testid="stDateInput"] input,
    div[data-testid="stTextInput"] input {
        background-color: rgba(15, 23, 42, 0.96) !important;
        color: #ffffff !important;
        border-color: rgba(56, 189, 248, 0.30) !important;
        border-radius: 12px !important;
    }

    div[data-baseweb="select"] span,
    div[data-baseweb="select"] input {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
    }

    div[data-baseweb="tag"] {
        background-color: rgba(14, 165, 233, 0.22) !important;
        border: 1px solid rgba(56, 189, 248, 0.35) !important;
    }

    div[data-baseweb="tag"] span {
        color: #e0f2fe !important;
    }

    button[kind="primary"], button[kind="secondary"] {
        border-radius: 12px !important;
        border: 1px solid rgba(56, 189, 248, 0.30) !important;
        background: rgba(15, 23, 42, 0.85) !important;
        color: #ffffff !important;
    }


    /* KOMPAKTAN LAYOUT ZA ONLINE VERZIJU */
    .block-container {
        padding-top: 0.6rem !important;
        padding-left: 1.2rem !important;
        padding-right: 1.2rem !important;
    }

    .hero-title {
        width: 100%;
        text-align: center;
        font-size: clamp(30px, 3.1vw, 52px);
        line-height: 1.15;
        font-weight: 1000;
        letter-spacing: .2px;
        color: #ffffff !important;
        text-shadow: 0 0 22px rgba(56,189,248,.35);
        margin: 4px 0 12px 0;
        padding: 8px 0 2px 0;
        white-space: normal;
        overflow: visible;
    }

    .toolbar-card {
        background: rgba(2, 6, 23, .70);
        border: 1px solid rgba(56,189,248,.22);
        border-radius: 14px;
        padding: 8px 10px;
        margin: 4px 0 10px 0;
        box-shadow: 0 12px 30px rgba(0,0,0,.24);
        backdrop-filter: blur(10px);
    }

    .toolbar-card label, .toolbar-card p {
        font-size: 12px !important;
        margin-bottom: 2px !important;
        color: #bfdbfe !important;
        font-weight: 800 !important;
    }

    .toolbar-card [data-testid="stFileUploader"] section {
        min-height: 46px !important;
        padding: 4px 8px !important;
        border-radius: 12px !important;
    }

    .toolbar-card [data-testid="stFileUploader"] section > div {
        padding: 0 !important;
    }

    .toolbar-card [data-testid="stFileUploader"] small,
    .toolbar-card [data-testid="stFileUploader"] button,
    .toolbar-card [data-testid="stFileUploaderDropzoneInstructions"] {
        font-size: 11px !important;
    }

    .toolbar-card div[data-baseweb="select"] > div,
    .toolbar-card div[data-testid="stDateInput"] input,
    .toolbar-card button[kind="secondary"] {
        min-height: 40px !important;
        height: 40px !important;
    }

    /* Sidebar uži i uvek pregledan */
    section[data-testid="stSidebar"] {
        min-width: 205px !important;
        max-width: 220px !important;
        width: 215px !important;
        background: rgba(2,6,23,.96) !important;
        border-right: 1px solid rgba(56,189,248,.22) !important;
    }

    /* Sidebar mora da ostane otvoren: sakrij dugmad za skupljanje/otvaranje menija */
    [data-testid="collapsedControl"],
    [data-testid="stSidebarCollapseButton"],
    button[data-testid="stSidebarCollapseButton"],
    button[title="Hide sidebar"],
    button[title="Show sidebar"] {
        display: none !important;
        visibility: hidden !important;
        pointer-events: none !important;
    }

    section[data-testid="stSidebar"] {
        transform: none !important;
        visibility: visible !important;
    }


    section[data-testid="stSidebar"] [role="radiogroup"] label {
        border: 1px solid rgba(56,189,248,.18) !important;
        border-radius: 12px !important;
        margin-bottom: 7px !important;
        padding: 7px 9px !important;
        background: rgba(15,23,42,.62) !important;
    }

    /* Ako se negde otvori native kalendar, tekst neka bude taman i čitljiv */
    div[data-baseweb="calendar"], div[data-baseweb="calendar"] * {
        color: #0f172a !important;
        -webkit-text-fill-color: #0f172a !important;
    }

    div[data-baseweb="calendar"] button {
        color: #0f172a !important;
        -webkit-text-fill-color: #0f172a !important;
    }

    /* Popover kalendar za izbor pojedinačnih dana */
    div[data-testid="stPopover"] button {
        background: rgba(15,23,42,.98) !important;
        color: #ffffff !important;
        border: 1px solid rgba(56,189,248,.28) !important;
        border-radius: 12px !important;
    }

    /* FIX V4: levi meni bez duplog skrola.
       Ostaje otvoren, sužen i bez unutrašnjih scrollbar-ova. */
    section[data-testid="stSidebar"] {
        min-width: 190px !important;
        max-width: 205px !important;
        width: 200px !important;
        overflow: hidden !important;
        height: 100vh !important;
        max-height: 100vh !important;
    }

    section[data-testid="stSidebar"] > div,
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        overflow: hidden !important;
        height: auto !important;
        max-height: none !important;
        padding-bottom: .6rem !important;
    }

    section[data-testid="stSidebar"]::-webkit-scrollbar,
    section[data-testid="stSidebar"] > div::-webkit-scrollbar,
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"]::-webkit-scrollbar {
        display: none !important;
        width: 0 !important;
    }

    section[data-testid="stSidebar"] h3 {
        font-size: 15px !important;
        margin-bottom: 6px !important;
    }

    section[data-testid="stSidebar"] [role="radiogroup"] label {
        padding: 5px 8px !important;
        margin-bottom: 5px !important;
        min-height: 34px !important;
    }

    section[data-testid="stSidebar"] [role="radiogroup"] label p {
        font-size: 13px !important;
        line-height: 1.15 !important;
        color: #f8fafc !important;
        -webkit-text-fill-color: #f8fafc !important;
        font-weight: 700 !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"] [role="radiogroup"] label,
    section[data-testid="stSidebar"] [role="radiogroup"] label * {
        color: #f8fafc !important;
        -webkit-text-fill-color: #f8fafc !important;
        opacity: 1 !important;
    }

    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
        background: linear-gradient(135deg, rgba(14,165,233,.30), rgba(37,99,235,.24)) !important;
        border-color: rgba(125,211,252,.75) !important;
        box-shadow: 0 0 0 1px rgba(56,189,248,.16), 0 8px 22px rgba(2,132,199,.14) !important;
    }

    /* Popover meniji ostaju kompaktni nakon odabira više stavki. */
    div[data-testid="stPopover"] > button {
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
    }

    </style>
    """,
    unsafe_allow_html=True
)


# ============================================================
# STITCH INDUSTRIAL PRECISION — FINALNI VIZUELNI OVERRIDE
# ============================================================
st.markdown(
    """
    <style>
    :root {
        --surface: #131313;
        --surface-lowest: #0e0e0e;
        --surface-low: #1c1b1b;
        --surface-container: #201f1f;
        --surface-high: #2a2a2a;
        --surface-highest: #353534;
        --on-surface: #e5e2e1;
        --on-surface-variant: #bbcabf;
        --outline: #86948a;
        --outline-variant: #3c4a42;
        --primary: #4edea3;
        --primary-container: #10b981;
        --secondary: #adc6ff;
        --tertiary: #ffb95f;
        --error: #ffb4ab;
    }

    html, body, [class*="css"] {
        font-family: Inter, "Segoe UI", Arial, sans-serif !important;
    }

    .stApp {
        background: var(--surface) !important;
        color: var(--on-surface) !important;
    }

    header[data-testid="stHeader"] {
        background: var(--surface-container) !important;
        border-bottom: 1px solid var(--outline-variant) !important;
        height: 56px !important;
    }

    .block-container {
        padding: 18px 24px 36px 24px !important;
        max-width: 100% !important;
    }

    .hero-title {
        margin: -4px 0 18px 0 !important;
        padding: 0 !important;
        text-align: left !important;
        color: var(--primary) !important;
        font-size: 32px !important;
        line-height: 34px !important;
        font-weight: 700 !important;
        letter-spacing: -0.02em !important;
        text-shadow: none !important;
        width: auto !important;
    }

    section[data-testid="stSidebar"] {
        width: 264px !important;
        min-width: 264px !important;
        max-width: 264px !important;
        background: var(--surface-low) !important;
        border-right: 1px solid var(--outline-variant) !important;
        overflow: hidden !important;
        height: 100vh !important;
        max-height: 100vh !important;
    }

    section[data-testid="stSidebar"] > div,
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        overflow: hidden !important;
        height: auto !important;
        max-height: none !important;
        padding: 0 !important;
    }

    section[data-testid="stSidebar"]::-webkit-scrollbar,
    section[data-testid="stSidebar"] > div::-webkit-scrollbar,
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"]::-webkit-scrollbar {
        display: none !important;
        width: 0 !important;
    }

    .sidebar-brand {
        padding: 24px 24px 20px 24px;
        border-bottom: 1px solid transparent;
    }
    .sidebar-brand-title {
        color: var(--primary);
        font-size: 20px;
        line-height: 28px;
        font-weight: 800;
        letter-spacing: -.01em;
    }
    .sidebar-brand-sub {
        color: var(--on-surface-variant);
        font-size: 14px;
        line-height: 20px;
        opacity: .72;
    }
    .sidebar-footer {
        margin-top: 28px;
        padding: 18px 24px;
        border-top: 1px solid var(--outline-variant);
        color: var(--on-surface-variant);
        font-size: 11px;
        font-weight: 700;
        letter-spacing: .08em;
    }

    section[data-testid="stSidebar"] [role="radiogroup"] {
        padding: 6px 12px 10px 12px !important;
        gap: 0 !important;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label {
        min-height: 52px !important;
        padding: 12px 14px !important;
        margin: 0 0 2px 0 !important;
        border: 0 !important;
        border-left: 4px solid transparent !important;
        border-radius: 0 !important;
        background: transparent !important;
        box-shadow: none !important;
        transition: background .15s ease, border-color .15s ease !important;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label:hover {
        background: var(--surface-container) !important;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
        background: var(--surface-high) !important;
        border-left-color: var(--primary) !important;
        box-shadow: none !important;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label p,
    section[data-testid="stSidebar"] [role="radiogroup"] label span,
    section[data-testid="stSidebar"] [role="radiogroup"] label * {
        color: var(--on-surface-variant) !important;
        -webkit-text-fill-color: var(--on-surface-variant) !important;
        font-size: 15px !important;
        line-height: 20px !important;
        font-weight: 500 !important;
        opacity: 1 !important;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) p,
    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) span,
    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) * {
        color: var(--primary) !important;
        -webkit-text-fill-color: var(--primary) !important;
        font-weight: 700 !important;
    }

    .toolbar-card {
        background: var(--surface-container) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
        padding: 10px 12px !important;
        margin: 0 0 16px 0 !important;
        box-shadow: none !important;
        backdrop-filter: none !important;
    }
    .toolbar-card label, .toolbar-card p {
        color: var(--on-surface-variant) !important;
        font-size: 12px !important;
        font-weight: 600 !important;
        letter-spacing: .04em !important;
        text-transform: uppercase !important;
    }

    .toolbar-mini-label {
        color: var(--on-surface-variant) !important;
        font-size: 12px !important;
        font-weight: 600 !important;
        letter-spacing: .04em !important;
        text-transform: uppercase !important;
        margin: 0 0 6px 2px !important;
        line-height: 1 !important;
    }

    [data-testid="stFileUploader"] section {
        min-height: 52px !important;
        padding: 4px 10px !important;
        background: var(--surface) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
    }
    [data-testid="stFileUploader"] section:hover {
        border-color: var(--primary) !important;
    }

    div[data-baseweb="select"] > div,
    div[data-testid="stDateInput"] input,
    div[data-testid="stTextInput"] input,
    div[data-testid="stPopover"] > button,
    button[kind="secondary"] {
        min-height: 44px !important;
        background: var(--surface) !important;
        color: var(--on-surface) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
        box-shadow: none !important;
    }
    div[data-baseweb="select"] > div:hover,
    div[data-testid="stPopover"] > button:hover,
    button[kind="secondary"]:hover {
        border-color: var(--primary) !important;
    }
    div[data-baseweb="select"] span,
    div[data-baseweb="select"] input,
    div[data-testid="stPopover"] button,
    div[data-testid="stPopover"] button * {
        color: var(--on-surface) !important;
        -webkit-text-fill-color: var(--on-surface) !important;
    }
    div[role="listbox"], ul[role="listbox"] {
        background: var(--surface-high) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
    }
    div[role="option"], li[role="option"] {
        color: var(--on-surface) !important;
        background: transparent !important;
    }
    div[role="option"]:hover, li[role="option"]:hover,
    div[aria-selected="true"][role="option"] {
        background: var(--surface-highest) !important;
        color: var(--primary) !important;
    }

    .neo-card,
    .reason-card,
    div[data-testid="stMetric"],
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: var(--surface-container) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
        box-shadow: none !important;
        backdrop-filter: none !important;
    }
    .neo-card {
        min-height: 132px !important;
        padding: 18px 20px !important;
        border-left: 3px solid var(--primary) !important;
    }
    .neo-card-green { border-left-color: var(--primary) !important; }
    .neo-card-orange { border-left-color: var(--tertiary) !important; }
    .neo-card-purple { border-left-color: var(--error) !important; }
    .metric-label-modern {
        color: var(--on-surface-variant) !important;
        font-size: 12px !important;
        line-height: 16px !important;
        font-weight: 600 !important;
        letter-spacing: .05em !important;
        text-transform: uppercase !important;
    }
    .metric-value-modern {
        color: var(--on-surface) !important;
        font-size: clamp(32px, 3vw, 48px) !important;
        line-height: 56px !important;
        font-weight: 700 !important;
        letter-spacing: -.02em !important;
    }
    .metric-sub-modern {
        color: var(--primary) !important;
        font-size: 13px !important;
        line-height: 18px !important;
        font-weight: 600 !important;
    }

    h1, h2, h3, h4, [data-testid="stHeading"] {
        color: var(--on-surface) !important;
        text-shadow: none !important;
    }
    h2, h3 {
        font-weight: 600 !important;
        letter-spacing: -.01em !important;
    }
    p, span, label, div {
        text-rendering: optimizeLegibility;
    }

    div[data-testid="stDataFrame"] {
        background: var(--surface-container) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
        box-shadow: none !important;
    }

    .js-plotly-plot, .plot-container, .svg-container {
        border-radius: 4px !important;
    }
    [data-testid="stPlotlyChart"] {
        background: var(--surface-container) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
        padding: 8px !important;
    }

    [data-testid="stAlert"] {
        background: var(--surface-container) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
        color: var(--on-surface) !important;
    }

    div[data-testid="stPopoverBody"] {
        background: var(--surface-high) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
        color: var(--on-surface) !important;
    }

    div[data-baseweb="calendar"] {
        background: var(--surface-high) !important;
        border: 1px solid var(--outline-variant) !important;
        border-radius: 4px !important;
    }
    div[data-baseweb="calendar"], div[data-baseweb="calendar"] * {
        color: var(--on-surface) !important;
        -webkit-text-fill-color: var(--on-surface) !important;
    }
    div[data-baseweb="calendar"] button:hover {
        background: var(--surface-highest) !important;
    }

    button[kind="primary"] {
        background: var(--primary) !important;
        color: #003824 !important;
        border: 1px solid var(--primary) !important;
        border-radius: 4px !important;
        font-weight: 800 !important;
        box-shadow: none !important;
    }
    button[kind="primary"] * {
        color: #003824 !important;
    }

    hr {
        border-color: var(--outline-variant) !important;
    }


    section[data-testid="stSidebar"] [role="radiogroup"] label {
        border: 1px dashed var(--outline-variant) !important;
        border-left: 1px dashed var(--outline-variant) !important;
        border-radius: 6px !important;
        margin: 0 12px 8px 12px !important;
        min-height: 42px !important;
        padding: 9px 12px !important;
        background: transparent !important;
        position: relative !important;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
        background: var(--surface-high) !important;
        border-color: var(--primary) !important;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label > div:first-child {
        display: none !important;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label p::before {
        content: "";
        display: inline-block;
        width: 9px;
        height: 9px;
        border-radius: 50%;
        margin-right: 10px;
        vertical-align: 1px;
        background: #4edea3;
        box-shadow: 0 0 8px currentColor;
    }
    section[data-testid="stSidebar"] [role="radiogroup"] label:nth-child(2) p::before { background:#ffb4ab; }
    section[data-testid="stSidebar"] [role="radiogroup"] label:nth-child(3) p::before { background:#ffb95f; }
    section[data-testid="stSidebar"] [role="radiogroup"] label:nth-child(4) p::before { background:#adc6ff; }
    section[data-testid="stSidebar"] [role="radiogroup"] label:nth-child(5) p::before { background:#c6b6ff; }
    section[data-testid="stSidebar"] [role="radiogroup"] label:nth-child(6) p::before { background:#7dd3fc; }
    section[data-testid="stSidebar"] [role="radiogroup"] label:nth-child(7) p::before { background:#f9a8d4; }
    section[data-testid="stSidebar"] [role="radiogroup"] label:nth-child(8) p::before { background:#fde68a; }


    /* Naslovi grafikona i kartica: levo poravnati i bez odsecanja */
    .neo-card,
    .reason-card,
    .summary-card,
    .machine-card {
        text-align: left !important;
    }
    .metric-label-modern,
    .metric-value-modern,
    .metric-sub-modern,
    .machine-title,
    .reason-title,
    .reason-sub,
    h1, h2, h3, h4,
    [data-testid="stHeading"] {
        text-align: left !important;
        white-space: normal !important;
        overflow: visible !important;
        text-overflow: clip !important;
        word-break: normal !important;
    }
    [data-testid="stPlotlyChart"] {
        overflow: visible !important;
    }

    @media (max-width: 1024px) {
        section[data-testid="stSidebar"] {
            width: 230px !important;
            min-width: 230px !important;
            max-width: 230px !important;
        }
        .block-container { padding: 16px !important; }
        .hero-title { font-size: 26px !important; line-height: 30px !important; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# Sakrij tekstualni title iz starog dela ako ga Streamlit prikaže previsoko.
st.markdown(
    """
    <div class="hero-title">Proizvodnja<br>dashboard</div>
    """,
    unsafe_allow_html=True
)


def _fmt_num(x):
    return format_broj(x)


def _fmt_pct_local(x):
    return format_proc(x)


def _safe_sum(df_obj, col):
    if df_obj is None or df_obj.empty or col not in df_obj.columns:
        return 0
    return pd.to_numeric(df_obj[col], errors="coerce").fillna(0).sum()


def _summary_realization_sources(df_obj, selected_processes):
    """Za gornje KPI kartice:
    - ako je izabran tačno jedan proces, i stator i rotor realizacija se prikazuju za taj proces
    - inače stator ide iz DMC, a rotor iz ROTOR (finalni procesi)
    """
    if df_obj is None or df_obj.empty:
        return pd.DataFrame(), pd.DataFrame(), "DMC finalni proces", "ROTOR finalni proces"

    if selected_processes and len(selected_processes) == 1:
        proc = selected_processes[0]
        filt = df_obj[df_obj["Proces"] == proc].copy()
        return filt, filt, f"{proc} proces", f"{proc} proces"

    df_stator = df_obj[df_obj["Proces"] == "DMC"].copy()
    df_rotor = df_obj[df_obj["Proces"] == "ROTOR"].copy()
    return df_stator, df_rotor, "DMC finalni proces", "ROTOR finalni proces"


def _dark_fig(fig, title=None, height=420):
    fig.update_layout(
        title=dict(text=title or "", x=0.01, xanchor="left", font=dict(size=22, color="#F8FAFC")),
        height=height,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#201f1f",
        font=dict(color="#e5e2e1"),
        margin=dict(l=35, r=25, t=70 if title else 25, b=55),
        hoverlabel=dict(bgcolor="#2a2a2a", font_color="#e5e2e1", bordercolor="#4edea3"),
    )
    fig.update_xaxes(gridcolor="#3c4a42", zerolinecolor="#3c4a42")
    fig.update_yaxes(gridcolor="#3c4a42", zerolinecolor="#3c4a42")
    return fig


def _html_escape(x):
    import html
    return html.escape(str(x))


def prikazi_metric_card(label, value, sub="", cls=""):
    st.markdown(
        f"""
        <div class="neo-card {cls}">
            <div class="metric-label-modern">{_html_escape(label)}</div>
            <div class="metric-value-modern">{_html_escape(value)}</div>
            <div class="metric-sub-modern">{_html_escape(sub)}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


def prikazi_dark_dataframe(df_show, use_container_width=True):
    if df_show is None or df_show.empty:
        st.info("Nema podataka za prikaz.")
        return
    # Pandas Styler daje tamniji prikaz i u Streamlit-u smanjuje efekat bele tabele.
    styled = (
        df_show.style
        .set_table_styles([
            {"selector": "thead th", "props": [("background-color", "#0f172a"), ("color", "#f8fafc"), ("border", "1px solid #334155")]},
            {"selector": "tbody td", "props": [("background-color", "#020617"), ("color", "#e5e7eb"), ("border", "1px solid #1e293b")]},
            {"selector": "table", "props": [("border-collapse", "separate"), ("border-spacing", "0"), ("border", "1px solid #334155")]},
        ])
    )
    st.dataframe(styled, use_container_width=use_container_width)


def obrisi_nemapirano(df_obj):
    if df_obj is None or df_obj.empty:
        return df_obj
    df_obj = df_obj.copy()
    for kol in ["Projekat", "Proces"]:
        if kol in df_obj.columns:
            df_obj[kol] = df_obj[kol].replace({"NEMAPIRANO": "OSTALO", None: "OSTALO"}).fillna("OSTALO")
    return df_obj


def prikazi_gauge_dashboard(naslov, df_kpi):
    kpi = izracunaj_kpi(df_kpi)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        prikazi_metric_card("Availability", _fmt_pct_local(kpi["Availability_pct"]), "raspoloživost", "")
    with c2:
        prikazi_metric_card("Realizacija", _fmt_pct_local(kpi["Realizacija_pct"]), f"kom: {_fmt_num(kpi['Realizacija'])}", "neo-card-green")
    with c3:
        prikazi_metric_card("Scrap rate", _fmt_pct_local(kpi["Scrap_rate_pct"]), f"NOK: {_fmt_num(kpi['NOK'])}", "neo-card-orange")
    with c4:
        prikazi_metric_card("OEE", _fmt_pct_local(kpi["OEE_pct"]), naslov, "neo-card-purple")

    vrednosti = [
        0 if kpi["Availability_pct"] is None else float(kpi["Availability_pct"]),
        0 if kpi["Realizacija_pct"] is None else float(kpi["Realizacija_pct"]),
        0 if kpi["Quality_pct"] is None else float(kpi["Quality_pct"]),
        0 if kpi["OEE_pct"] is None else float(kpi["OEE_pct"]),
    ]
    labels = ["Availability", "Realizacija", "Quality", "OEE"]
    fig = go.Figure()
    domains = [([0.04, 0.46], [0.55, 0.94]), ([0.54, 0.96], [0.55, 0.94]), ([0.04, 0.46], [0.04, 0.43]), ([0.54, 0.96], [0.04, 0.43])]
    for label, val, (xd, yd) in zip(labels, vrednosti, domains):
        fig.add_trace(go.Indicator(
            mode="gauge+number",
            value=val,
            number=dict(suffix="%", font=dict(size=30, color="#F8FAFC")),
            title=dict(text=label, font=dict(size=16, color="#CBD5E1")),
            domain=dict(x=xd, y=yd),
            gauge=dict(
                axis=dict(range=[0, max(110, val * 1.15)], tickcolor="#94A3B8"),
                bar=dict(color="#38BDF8"),
                bgcolor="rgba(15,23,42,0.75)",
                borderwidth=1,
                bordercolor="rgba(56,189,248,0.25)",
                steps=[
                    dict(range=[0, 70], color="rgba(239,68,68,0.18)"),
                    dict(range=[70, 90], color="rgba(249,115,22,0.18)"),
                    dict(range=[90, 100], color="rgba(59,130,246,0.18)"),
                    dict(range=[100, max(110, val * 1.15)], color="rgba(34,197,94,0.18)"),
                ],
                threshold=dict(line=dict(color="#F8FAFC", width=3), thickness=0.75, value=100),
            )
        ))
    st.plotly_chart(_dark_fig(fig, f"KPI globalni — {naslov}", height=560), use_container_width=True, config={"displayModeBar": False})


# ============================================================
# LEVI MENI - SAMO TABOVI/SEKCIJE
# ============================================================

sekcije = [
    "Dnevni pregled",
    "NOK razlozi",
    "Zastoji",
    "KPI",
    "Grafički prikaz",
    "SCRAP",
    "Top uzroci po mašini",
    "Tabele",
]

st.sidebar.markdown("""
<div class="sidebar-brand">
  <div class="sidebar-brand-title">Production monitoring</div>
  <div class="sidebar-brand-sub">Manufacturing Suite</div>
</div>
""", unsafe_allow_html=True)
aktivna_sekcija = st.sidebar.radio("Izaberi tab", sekcije, label_visibility="collapsed")
st.sidebar.markdown("<div class='sidebar-footer'>PRODUCTION ANALYTICS</div>", unsafe_allow_html=True)

if "overlay_slika" not in st.session_state:
    st.session_state.overlay_slika = None

# ============================================================
# TOP TOOLBAR - UPLOAD + FILTERI
# ============================================================

st.markdown('<div class="toolbar-card">', unsafe_allow_html=True)
up_col, refresh_col = st.columns([5.5, 0.55])
with up_col:
    uploaded_file = st.file_uploader(
        "Excel fajl",
        type=["xlsx"],
        help="Učitaj mesečni Production realization Excel fajl.",
        label_visibility="collapsed",
    )
with refresh_col:
    if st.button("🔄", help="Osveži / očisti cache", use_container_width=True):
        st.cache_data.clear()
        st.rerun()
st.markdown('</div>', unsafe_allow_html=True)

if uploaded_file is None:
    st.info("Učitaj Excel fajl da bi se prikazao dashboard.")
    st.stop()

# Snimamo upload u privremeni fajl, da postojeća funkcija ucitaj_sve_podatke može da radi bez izmene logike čitanja.
bytes_data = uploaded_file.getvalue()
fajl_hash = hashlib.md5(bytes_data).hexdigest()[:12]
safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", uploaded_file.name)
privremeni_fajl = os.path.join(tempfile.gettempdir(), f"proizvodnja_{fajl_hash}_{safe_name}")
with open(privremeni_fajl, "wb") as f:
    f.write(bytes_data)

try:
    df, df_nok, df_zastoji = ucitaj_sve_podatke(privremeni_fajl)
except Exception as e:
    st.error("Greška pri učitavanju podataka.")
    st.code(str(e))
    st.stop()

if df.empty:
    st.warning("Nema podataka za prikaz.")
    st.stop()

# Projekat/proces + eliminacija teksta NEMAPIRANO iz prikaza.
df["Projekat"] = df["Masina"].apply(projekat_iz_masine)
df["Proces"] = df["Masina"].apply(proces_iz_masine)
df = obrisi_nemapirano(df)

if not df_nok.empty:
    df_nok["Projekat"] = df_nok["Masina"].apply(projekat_iz_masine)
    df_nok["Proces"] = df_nok["Masina"].apply(proces_iz_masine)
    df_nok = obrisi_nemapirano(df_nok)

if not df_zastoji.empty:
    df_zastoji["Projekat"] = df_zastoji["Masina"].apply(projekat_iz_masine)
    df_zastoji["Proces"] = df_zastoji["Masina"].apply(proces_iz_masine)
    df_zastoji = obrisi_nemapirano(df_zastoji)

# Toolbar filteri se prikazuju tek kada znamo opcije iz fajla.
df_valid = df.dropna(subset=["Datum"]).copy()
svi_datumi = sorted(df_valid["Datum"].dt.date.unique())

redosled_projekata = ["APP350", "APP550", "VITESKO EMR4", "LK-4", "OSTALO"]
svi_projekti = [p for p in redosled_projekata if p in df["Projekat"].dropna().unique()]
for p in sorted(df["Projekat"].dropna().unique()):
    if p not in svi_projekti:
        svi_projekti.append(p)

redosled_procesa = ["STAMPING", "WELDING", "DMC", "GP12", "ROTOR", "OSTALO"]
svi_procesi = [p for p in redosled_procesa if p in df["Proces"].dropna().unique()]
for p in sorted(df["Proces"].dropna().unique()):
    if p not in svi_procesi:
        svi_procesi.append(p)

st.markdown('<div class="toolbar-card">', unsafe_allow_html=True)
f1, f2, f3, f4, f5 = st.columns([1.15, 1.05, 1.05, 1.55, 0.95])

# --- DATUM: kompaktan popover kalendar za izbor pojedinačnih dana ---
datum_key = f"izabrani_datumi_{fajl_hash}"
if datum_key not in st.session_state:
    st.session_state[datum_key] = [svi_datumi[-1].isoformat()] if svi_datumi else []

# Ukloni datume koji ne postoje u trenutno učitanom fajlu.
dostupni_iso = {d.isoformat() for d in svi_datumi}
st.session_state[datum_key] = [d for d in st.session_state[datum_key] if d in dostupni_iso]

with f1:
    broj_dana = len(st.session_state[datum_key])
    natpis = "Svi datumi" if broj_dana == 0 else f"Datum · {broj_dana} dan(a)"
    with st.popover(natpis, use_container_width=True):
        st.caption("Klikni na dane koje želiš da prikažeš. Ako ništa nije označeno, prikazuju se svi dani.")
        if svi_datumi:
            meseci = sorted({date(d.year, d.month, 1) for d in svi_datumi})
            oznake_meseci = [m.strftime("%m.%Y") for m in meseci]
            podrazumevani_mesec = max(0, len(meseci) - 1)
            izabrani_mesec_oznaka = st.selectbox("Mesec", oznake_meseci, index=podrazumevani_mesec)
            izabrani_mesec = meseci[oznake_meseci.index(izabrani_mesec_oznaka)]

            c_svi, c_ocisti = st.columns(2)
            # Važno: dugme označava samo dane iz trenutno izabranog meseca,
            # a ne sve datume iz celog Excel fajla.
            datumi_iz_meseca = sorted(
                d.isoformat()
                for d in svi_datumi
                if d.year == izabrani_mesec.year and d.month == izabrani_mesec.month
            )

            c_svi, c_ocisti = st.columns(2)
            with c_svi:
                if st.button("Označi dane iz meseca", use_container_width=True):
                    st.session_state[datum_key] = datumi_iz_meseca
                    st.rerun()
            with c_ocisti:
                if st.button("Očisti izbor", use_container_width=True):
                    st.session_state[datum_key] = []
                    st.rerun()

            st.caption(f"Označava se samo mesec {izabrani_mesec.strftime('%m.%Y')}.")
            st.markdown("**PON UTO SRE ČET PET SUB NED**")
            kal = calendar.Calendar(firstweekday=0)
            for nedelja in kal.monthdayscalendar(izabrani_mesec.year, izabrani_mesec.month):
                cols = st.columns(7)
                for i, dan in enumerate(nedelja):
                    if dan == 0:
                        cols[i].markdown(" ")
                        continue
                    d = date(izabrani_mesec.year, izabrani_mesec.month, dan)
                    d_iso = d.isoformat()
                    if d_iso not in dostupni_iso:
                        cols[i].button(str(dan), disabled=True, use_container_width=True, key=f"day_disabled_{d_iso}")
                    else:
                        oznacen = d_iso in st.session_state[datum_key]
                        label = f"✓ {dan}" if oznacen else str(dan)
                        if cols[i].button(label, use_container_width=True, key=f"day_toggle_{d_iso}"):
                            if oznacen:
                                st.session_state[datum_key].remove(d_iso)
                            else:
                                st.session_state[datum_key].append(d_iso)
                                st.session_state[datum_key] = sorted(set(st.session_state[datum_key]))
                            st.rerun()

izabrani_datumi = [date.fromisoformat(d) for d in st.session_state[datum_key] if d in dostupni_iso]

# Projekat i proces su kompaktni popover filteri sa višestrukim izborom.
# Nakon zatvaranja prikazuje se samo broj izabranih stavki, bez crvenih tagova.
projekat_key = f"projekti_filter_{fajl_hash}"
proces_key = f"procesi_filter_{fajl_hash}"

if projekat_key not in st.session_state:
    st.session_state[projekat_key] = list(svi_projekti)
else:
    st.session_state[projekat_key] = [p for p in st.session_state[projekat_key] if p in svi_projekti]

if proces_key not in st.session_state:
    st.session_state[proces_key] = list(svi_procesi)
else:
    st.session_state[proces_key] = [p for p in st.session_state[proces_key] if p in svi_procesi]

with f2:
    broj_p = len(st.session_state[projekat_key])
    naslov_p = "Projekat · SVI" if broj_p == len(svi_projekti) else f"Projekat · {broj_p}"
    with st.popover(naslov_p, use_container_width=True):
        c1, c2 = st.columns(2)
        if c1.button("Označi sve", key="projekti_svi", use_container_width=True):
            st.session_state[projekat_key] = list(svi_projekti)
            st.rerun()
        if c2.button("Očisti", key="projekti_ocisti", use_container_width=True):
            st.session_state[projekat_key] = []
            st.rerun()
        for p in svi_projekti:
            checked = p in st.session_state[projekat_key]
            novo = st.checkbox(p, value=checked, key=f"projekat_cb_{fajl_hash}_{p}")
            if novo != checked:
                if novo:
                    st.session_state[projekat_key].append(p)
                else:
                    st.session_state[projekat_key].remove(p)
                st.rerun()

with f3:
    broj_pr = len(st.session_state[proces_key])
    naslov_pr = "Proces · SVI" if broj_pr == len(svi_procesi) else f"Proces · {broj_pr}"
    with st.popover(naslov_pr, use_container_width=True):
        c1, c2 = st.columns(2)
        if c1.button("Označi sve", key="procesi_svi", use_container_width=True):
            st.session_state[proces_key] = list(svi_procesi)
            st.rerun()
        if c2.button("Očisti", key="procesi_ocisti", use_container_width=True):
            st.session_state[proces_key] = []
            st.rerun()
        for p in svi_procesi:
            checked = p in st.session_state[proces_key]
            novo = st.checkbox(p, value=checked, key=f"proces_cb_{fajl_hash}_{p}")
            if novo != checked:
                if novo:
                    st.session_state[proces_key].append(p)
                else:
                    st.session_state[proces_key].remove(p)
                st.rerun()

izabrani_projekti = list(st.session_state[projekat_key])
izabrani_procesi = list(st.session_state[proces_key])

# Mašine zavise od projekta/procesa.
df_za_masine = df.copy()
if izabrani_projekti:
    df_za_masine = df_za_masine[df_za_masine["Projekat"].isin(izabrani_projekti)]
if izabrani_procesi:
    df_za_masine = df_za_masine[df_za_masine["Proces"].isin(izabrani_procesi)]
sve_masine = sorted(df_za_masine["Masina"].dropna().unique())
with f4:
    st.markdown("<div class='toolbar-mini-label'>Mašina</div>", unsafe_allow_html=True)
    masina_filter = st.selectbox("Mašina", ["SVE"] + sve_masine, index=0, label_visibility="collapsed")
izabrane_masine = sve_masine if masina_filter == "SVE" else [masina_filter]

with f5:
    st.markdown("<div class='toolbar-mini-label'>Grafikon</div>", unsafe_allow_html=True)
    tip_grafikona = st.selectbox("Grafikon", ["Stubičasti", "Linijski", "Površinski"], index=0, label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)

# ============================================================
# PRIMENA FILTERA
# ============================================================

df_filter = df.copy()
if izabrani_datumi:
    df_filter = df_filter[df_filter["Datum"].dt.date.isin(izabrani_datumi)]
if izabrani_projekti:
    df_filter = df_filter[df_filter["Projekat"].isin(izabrani_projekti)]
if izabrani_procesi:
    df_filter = df_filter[df_filter["Proces"].isin(izabrani_procesi)]
if izabrane_masine:
    df_filter = df_filter[df_filter["Masina"].isin(izabrane_masine)]

df_ukupno_filter = df_filter[~df_filter["Masina"].isin(iskljuci_iz_ukupnog_proracuna)].copy()

df_nok_filter = finalno_ocisti_df_razloge(df_nok.copy())
df_zastoji_filter = finalno_ocisti_df_razloge(df_zastoji.copy())

if not df_nok_filter.empty:
    if izabrani_datumi:
        df_nok_filter = df_nok_filter[df_nok_filter["Datum"].dt.date.isin(izabrani_datumi)]
    if izabrani_projekti:
        df_nok_filter = df_nok_filter[df_nok_filter["Projekat"].isin(izabrani_projekti)]
    if izabrani_procesi:
        df_nok_filter = df_nok_filter[df_nok_filter["Proces"].isin(izabrani_procesi)]
    if izabrane_masine:
        df_nok_filter = df_nok_filter[df_nok_filter["Masina"].isin(izabrane_masine)]

if not df_zastoji_filter.empty:
    if izabrani_datumi:
        df_zastoji_filter = df_zastoji_filter[df_zastoji_filter["Datum"].dt.date.isin(izabrani_datumi)]
    if izabrani_projekti:
        df_zastoji_filter = df_zastoji_filter[df_zastoji_filter["Projekat"].isin(izabrani_projekti)]
    if izabrani_procesi:
        df_zastoji_filter = df_zastoji_filter[df_zastoji_filter["Proces"].isin(izabrani_procesi)]
    if izabrane_masine:
        df_zastoji_filter = df_zastoji_filter[df_zastoji_filter["Masina"].isin(izabrane_masine)]

if st.session_state.get("overlay_slika") is not None:
    prikazi_overlay_sliku(st.session_state.overlay_slika["putanja"], st.session_state.overlay_slika["naslov"])

# Globalni brzi KPI iznad sekcija.
df_real_stator, df_real_rotor, stator_sub, rotor_sub = _summary_realization_sources(df_ukupno_filter, izabrani_procesi)
col_a, col_b, col_c, col_d = st.columns(4)
with col_a:
    prikazi_metric_card("Realizacija STATOR", _fmt_num(_safe_sum(df_real_stator, "Realizacija_STATOR")), stator_sub, "")
with col_b:
    prikazi_metric_card("Realizacija ROTOR", _fmt_num(_safe_sum(df_real_rotor, "Realizacija_ROTOR")), rotor_sub, "neo-card-green")
with col_c:
    prikazi_metric_card("Zastoji", f"{_fmt_num(_safe_sum(df_zastoji_filter, 'Minuta_iz_note'))} min", "po notes-ima", "neo-card-orange")
with col_d:
    prikazi_metric_card("NOK", _fmt_num(_safe_sum(df_nok_filter, "Komada_iz_note")), "po notes-ima", "neo-card-purple")


def _bar_chart(df_plot, x, y, title, horizontal=False, color="#4edea3"):
    if df_plot is None or df_plot.empty:
        st.info("Nema podataka za prikaz.")
        return
    if horizontal:
        fig = go.Figure(go.Bar(x=df_plot[y], y=df_plot[x], orientation="h", text=df_plot[y].apply(_fmt_num), textposition="auto", marker_color=color))
        fig.update_yaxes(autorange="reversed")
    else:
        fig = go.Figure(go.Bar(x=df_plot[x], y=df_plot[y], text=df_plot[y].apply(_fmt_num), textposition="outside", marker_color=color))
    st.plotly_chart(_dark_fig(fig, title), use_container_width=True, config={"displayModeBar": False})


def _ok_nok_chart(ok, nok, title):
    d = pd.DataFrame({"Tip":["OK","NOK"], "Vrednost":[ok,nok]})
    fig = go.Figure(go.Bar(x=d["Tip"], y=d["Vrednost"], text=d["Vrednost"].apply(_fmt_num), textposition="outside", marker_color=["#4edea3","#ffb4ab"]))
    st.plotly_chart(_dark_fig(fig, title), use_container_width=True, config={"displayModeBar": False})


def _pareto(df_src, category, value, title):
    if df_src is None or df_src.empty:
        st.info("Nema podataka za Pareto prikaz.")
        return
    p = df_src.groupby(category, as_index=False)[value].sum().sort_values(value, ascending=False)
    p = p[p[value] > 0].head(12).copy()
    if p.empty:
        st.info("Nema pozitivnih vrednosti za Pareto prikaz.")
        return
    total = p[value].sum()
    p["Kumulativno"] = p[value].cumsum()/total*100
    fig = go.Figure()
    fig.add_trace(go.Bar(x=p[category], y=p[value], name=value, text=p[value].apply(_fmt_num), textposition="outside", marker_color="#4edea3"))
    fig.add_trace(go.Scatter(x=p[category], y=p["Kumulativno"], name="Kumulativno %", yaxis="y2", mode="lines+markers", line=dict(color="#ffb95f", width=3)))
    fig.update_layout(yaxis2=dict(overlaying="y", side="right", range=[0,110], ticksuffix="%"), legend=dict(orientation="h"))
    st.plotly_chart(_dark_fig(fig, title, height=500), use_container_width=True, config={"displayModeBar": False})
    prikazi_dark_dataframe(p.rename(columns={value:"Vrednost", "Kumulativno":"Kumulativno %"}))



def _machine_header_card(name, total_minutes, color_index=0):
    colors = ["#4edea3", "#ffb95f", "#adc6ff", "#ffb4ab", "#9b8cff", "#64d8cb"]
    accent = colors[color_index % len(colors)]
    st.markdown(
        f"""
        <div style="border:1px solid {accent}66;border-radius:18px;padding:14px 16px;margin-bottom:12px;background:linear-gradient(135deg,{accent}22,rgba(20,25,35,.92));">
            <div style="font-size:22px;font-weight:800;color:#f8fafc;">{_html_escape(name)}</div>
        </div>
        <div style="border:1px solid #334155;border-radius:14px;padding:14px 16px;margin-bottom:12px;background:rgba(2,6,23,.90);">
            <div style="font-size:14px;color:#cbd5e1;font-weight:700;">Ukupno minuta zastoja</div>
            <div style="font-size:30px;color:#fff;font-weight:900;margin-top:6px;">{_fmt_num(total_minutes)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _render_machine_detail_grid(dfz):
    machines = sorted(dfz["Masina"].dropna().unique())
    for start in range(0, len(machines), 3):
        cols = st.columns(3)
        for offset, machine in enumerate(machines[start:start+3]):
            with cols[offset]:
                m = dfz[dfz["Masina"] == machine].copy()
                _machine_header_card(machine, _safe_sum(m, "Minuta_iz_note"), start + offset)
                cols_show = [c for c in ["Datum", "Smena", "Minuta_iz_note", "Razlog", "Originalna_stavka"] if c in m.columns]
                detail = m[cols_show].copy()
                if "Datum" in detail.columns:
                    detail["Datum"] = detail["Datum"].dt.strftime("%d.%m.%Y")
                detail = detail.rename(columns={"Minuta_iz_note":"Min", "Originalna_stavka":"Originalni unos"})
                st.dataframe(detail, use_container_width=True, hide_index=True, height=310)


def _process_machine_summary(df_proc):
    if df_proc.empty:
        return pd.DataFrame()
    cols = ["Plan_STATOR","Realizacija_STATOR","OK_STATOR","NOK_STATOR","Plan_ROTOR","Realizacija_ROTOR","OK_ROTOR","NOK_ROTOR"]
    for c in cols:
        if c not in df_proc.columns:
            df_proc[c] = 0
    out = df_proc.groupby("Masina", as_index=False)[cols].sum()
    out["Plan_UKUPNO"] = out["Plan_STATOR"] + out["Plan_ROTOR"]
    out["Realizacija_UKUPNO"] = out["Realizacija_STATOR"] + out["Realizacija_ROTOR"]
    out["Realizacija_pct"] = out.apply(lambda r: procenat(r["Realizacija_UKUPNO"], r["Plan_UKUPNO"]), axis=1)
    return out


def _grouped_bar(df_plot, x, ys, title):
    if df_plot.empty:
        st.info("Nema podataka za prikaz.")
        return
    fig = go.Figure()
    palette = ["#4edea3", "#adc6ff", "#ffb95f", "#ffb4ab", "#9b8cff"]
    for i, y in enumerate(ys):
        if y in df_plot.columns:
            fig.add_trace(go.Bar(x=df_plot[x], y=df_plot[y], name=y.replace("_", " "), marker_color=palette[i % len(palette)]))
    fig.update_layout(barmode="group", xaxis_tickangle=-25)
    st.plotly_chart(_dark_fig(fig, title, height=480), use_container_width=True, config={"displayModeBar":False})


def _plan_vs_realization(df_plot, plan_col, real_col, title):
    if df_plot.empty:
        st.info("Nema podataka za prikaz.")
        return
    fig = go.Figure()
    fig.add_trace(go.Bar(x=df_plot["Masina"], y=df_plot[plan_col], name="Plan", marker_color="#86948a"))
    fig.add_trace(go.Bar(x=df_plot["Masina"], y=df_plot[real_col], name="Realizacija", marker_color="#4edea3"))
    fig.update_layout(barmode="group", xaxis_tickangle=-25)
    st.plotly_chart(_dark_fig(fig, title, height=460), use_container_width=True, config={"displayModeBar":False})


def _render_process_graphics(df_all, process, display_name):
    d = df_all[df_all["Proces"].astype(str).str.upper() == process.upper()].copy()
    if d.empty:
        return
    st.markdown(f"## {display_name}")
    s = _process_machine_summary(d)
    _grouped_bar(s, "Masina", ["Plan_STATOR","Realizacija_STATOR","OK_STATOR","NOK_STATOR"], f"{display_name} — STATOR po mašini")
    _plan_vs_realization(s, "Plan_STATOR", "Realizacija_STATOR", f"Realizacija u odnosu na plan — STATOR — {display_name}")
    _grouped_bar(s, "Masina", ["Plan_ROTOR","Realizacija_ROTOR","OK_ROTOR","NOK_ROTOR"], f"{display_name} — ROTOR po mašini")
    _plan_vs_realization(s, "Plan_ROTOR", "Realizacija_ROTOR", f"Realizacija u odnosu na plan — ROTOR — {display_name}")
    with st.expander(f"Prikaži tabelu — {display_name}"):
        t=s.copy(); t["Realizacija_pct"] = t["Realizacija_pct"].apply(format_proc)
        prikazi_dark_dataframe(t)


def _render_scrap_sections(df_all):
    d = df_all.copy()
    for c in ["OK_STATOR","OK_ROTOR","NOK_STATOR","NOK_ROTOR"]:
        if c not in d.columns: d[c]=0
        d[c] = pd.to_numeric(d[c], errors="coerce").fillna(0)
    d["OK"] = d["OK_STATOR"] + d["OK_ROTOR"]
    d["NOK"] = d["NOK_STATOR"] + d["NOK_ROTOR"]
    for group_col, label in [("Projekat","projektu"),("Proces","procesu")]:
        g=d.groupby(group_col, as_index=False)[["OK","NOK"]].sum()
        g["Scrap %"] = g.apply(lambda r: procenat(r["NOK"], r["OK"]+r["NOK"]), axis=1)
        st.markdown(f"### SCRAP po {label}")
        fig=go.Figure()
        fig.add_trace(go.Bar(x=g[group_col], y=g["NOK"], name="NOK", text=g["Scrap %"].apply(_fmt_pct_local), textposition="outside", marker_color="#ffb4ab"))
        st.plotly_chart(_dark_fig(fig, f"SCRAP po {label}", height=450), use_container_width=True, config={"displayModeBar":False})
        prikazi_dark_dataframe(g)


def _top3_horizontal(df_source, group_filter_col, group_value, title):
    d=df_source[df_source[group_filter_col]==group_value]
    if d.empty: return
    t=d.groupby("Razlog",as_index=False)["Minuta_iz_note"].sum().sort_values("Minuta_iz_note",ascending=False).head(3)
    fig=go.Figure(go.Bar(x=t["Minuta_iz_note"], y=t["Razlog"], orientation="h", text=t["Minuta_iz_note"].apply(_fmt_num), textposition="auto", marker_color="#4edea3"))
    fig.update_layout(yaxis=dict(autorange="reversed"))
    st.plotly_chart(_dark_fig(fig,title,height=300),use_container_width=True,config={"displayModeBar":False})


def _render_top_causes_by_process(dfz):
    for proc in sorted(dfz["Proces"].dropna().unique()):
        dp=dfz[dfz["Proces"]==proc]
        st.markdown(f"## {proc}")
        overall=dp.groupby("Razlog",as_index=False)["Minuta_iz_note"].sum().sort_values("Minuta_iz_note",ascending=False).head(3)
        fig=go.Figure(go.Bar(x=overall["Minuta_iz_note"],y=overall["Razlog"],orientation="h",text=overall["Minuta_iz_note"].apply(_fmt_num),textposition="auto",marker_color="#ffb95f"))
        fig.update_layout(yaxis=dict(autorange="reversed"))
        st.plotly_chart(_dark_fig(fig,f"Top 3 uzroka ukupno — {proc}",height=320),use_container_width=True,config={"displayModeBar":False})
        machines=sorted(dp["Masina"].dropna().unique())
        for start in range(0,len(machines),3):
            cols=st.columns(3)
            for i,m in enumerate(machines[start:start+3]):
                with cols[i]:
                    _top3_horizontal(dp,"Masina",m,f"Top 3 — {m}")

# ============================================================
# SEKCIJE
# ============================================================

if aktivna_sekcija == "Dnevni pregled":
    st.subheader("📌 Ukupan pregled proizvodnje")
    st.caption("Ako je izabran tačno jedan proces, realizacija prati taj proces. U ostalim slučajevima STATOR ide iz DMC, a ROTOR iz ROTOR procesa. NOK ostaje iz svih procesa.")

    df_stator_gotov = df_ukupno_filter[df_ukupno_filter["Proces"] == "DMC"].copy()
    df_rotor_gotov = df_ukupno_filter[df_ukupno_filter["Proces"] == "ROTOR"].copy()

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        prikazi_metric_card("OK STATOR", _fmt_num(_safe_sum(df_stator_gotov, "OK_STATOR")), "DMC finalni proces")
    with c2:
        prikazi_metric_card("OK ROTOR", _fmt_num(_safe_sum(df_rotor_gotov, "OK_ROTOR")), "ROTOR finalni proces", "neo-card-green")
    with c3:
        prikazi_metric_card("NOK STATOR", _fmt_num(_safe_sum(df_ukupno_filter, "NOK_STATOR")), "svi procesi", "neo-card-orange")
    with c4:
        prikazi_metric_card("NOK ROTOR", _fmt_num(_safe_sum(df_ukupno_filter, "NOK_ROTOR")), "svi procesi", "neo-card-purple")

    if not df_filter.empty:
        prikaz = df_filter.groupby(["Projekat", "Proces"], as_index=False).agg({
            "Realizacija_STATOR": "sum", "Realizacija_ROTOR": "sum", "OK_STATOR": "sum", "OK_ROTOR": "sum", "NOK_STATOR": "sum", "NOK_ROTOR": "sum", "Stops_min": "sum"
        })
        prikaz["Realizacija"] = prikaz["Realizacija_STATOR"] + prikaz["Realizacija_ROTOR"]
        fig = go.Figure(go.Bar(x=prikaz["Proces"], y=prikaz["Realizacija"], text=prikaz["Realizacija"].apply(lambda x: f"{x:,.0f}".replace(",", ".")), textposition="outside"))
        st.plotly_chart(_dark_fig(fig, "Realizacija po procesu"), use_container_width=True, config={"displayModeBar": False})
        prikazi_dark_dataframe(prikaz)

elif aktivna_sekcija == "NOK razlozi":
    st.subheader("NOK razlozi iz notes-a")
    if df_nok_filter.empty:
        st.info("Nema NOK notes podataka za izabrane filtere.")
    else:
        dfn = df_nok_filter.copy()
        dfn["Tip_proizvoda"] = dfn.get("Tip_proizvoda", "NEPOZNATO").fillna("NEPOZNATO")
        stator = dfn[dfn["Tip_proizvoda"].astype(str).str.contains("STATOR", case=False, na=False)]
        rotor = dfn[dfn["Tip_proizvoda"].astype(str).str.contains("ROTOR", case=False, na=False)]
        c1,c2,c3,c4=st.columns(4)
        with c1: prikazi_metric_card("NOK ukupno", _fmt_num(_safe_sum(dfn,"Komada_iz_note")), "iz notes-a")
        with c2: prikazi_metric_card("NOK STATOR", _fmt_num(_safe_sum(stator,"Komada_iz_note")), "iz notes-a")
        with c3: prikazi_metric_card("NOK ROTOR", _fmt_num(_safe_sum(rotor,"Komada_iz_note")), "iz notes-a")
        with c4: prikazi_metric_card("Broj razloga", str(dfn["Razlog"].nunique()), "različitih")

        st.markdown("### NOK razlozi odvojeno: STATOR / ROTOR")
        a,b=st.columns(2)
        with a:
            st.markdown("#### STATOR NOK razlozi")
            zs=stator.groupby("Razlog",as_index=False)["Komada_iz_note"].sum().sort_values("Komada_iz_note",ascending=False)
            _bar_chart(zs.head(15),"Razlog","Komada_iz_note","STATOR NOK razlozi",True,"#4edea3")
            prikazi_dark_dataframe(zs)
        with b:
            st.markdown("#### ROTOR NOK razlozi")
            zr=rotor.groupby("Razlog",as_index=False)["Komada_iz_note"].sum().sort_values("Komada_iz_note",ascending=False)
            _bar_chart(zr.head(15),"Razlog","Komada_iz_note","ROTOR NOK razlozi",True,"#adc6ff")
            prikazi_dark_dataframe(zr)

        st.markdown("### NOK po mašini, proizvodu i razlogu")
        det=dfn.groupby(["Masina","Tip_proizvoda","Razlog"],as_index=False)["Komada_iz_note"].sum().sort_values(["Masina","Tip_proizvoda","Komada_iz_note"],ascending=[True,True,False])
        prikazi_dark_dataframe(det)

        st.markdown("### NOK analiza po procesu")
        for proc in sorted(dfn["Proces"].dropna().unique()):
            base=df_ukupno_filter[df_ukupno_filter["Proces"]==proc]
            n=dfn[dfn["Proces"]==proc]
            ok=_safe_sum(base,"OK_STATOR")+_safe_sum(base,"OK_ROTOR")
            nok=_safe_sum(base,"NOK_STATOR")+_safe_sum(base,"NOK_ROTOR")
            l,r=st.columns(2)
            with l: _ok_nok_chart(ok,nok,f"Odnos OK i NOK — {proc}")
            with r:
                pm=n.groupby("Masina",as_index=False)["Komada_iz_note"].sum().sort_values("Komada_iz_note",ascending=False)
                _bar_chart(pm,"Masina","Komada_iz_note",f"NOK po mašini — {proc}",True,"#ffb4ab")

        st.markdown("### NOK analiza po projektu")
        for proj in sorted(dfn["Projekat"].dropna().unique()):
            base=df_ukupno_filter[df_ukupno_filter["Projekat"]==proj]
            n=dfn[dfn["Projekat"]==proj]
            ok=_safe_sum(base,"OK_STATOR")+_safe_sum(base,"OK_ROTOR")
            nok=_safe_sum(base,"NOK_STATOR")+_safe_sum(base,"NOK_ROTOR")
            l,r=st.columns(2)
            with l: _ok_nok_chart(ok,nok,f"Odnos OK i NOK — {proj}")
            with r:
                pm=n.groupby("Masina",as_index=False)["Komada_iz_note"].sum().sort_values("Komada_iz_note",ascending=False)
                _bar_chart(pm,"Masina","Komada_iz_note",f"NOK po mašini — {proj}",True,"#ffb95f")

        st.markdown("### Pareto NOK razloga")
        _pareto(dfn,"Razlog","Komada_iz_note","Pareto NOK razloga")
        with st.expander("Prikaži sve NOK notes stavke"):
            prikazi_dark_dataframe(dfn)

elif aktivna_sekcija == "Zastoji":
    st.subheader("Zastoji iz notes-a")
    if df_zastoji_filter.empty:
        st.info("Nema podataka o zastojima za izabrane filtere.")
    else:
        dfz=df_zastoji_filter.copy()
        c1,c2,c3=st.columns(3)
        with c1: prikazi_metric_card("Ukupno minuta", _fmt_num(_safe_sum(dfz,"Minuta_iz_note")), "zastoja")
        with c2: prikazi_metric_card("Broj razloga", str(dfz["Razlog"].nunique()), "različitih")
        with c3: prikazi_metric_card("Broj mašina", str(dfz["Masina"].nunique()), "sa zastojem")

        zbir=dfz.groupby("Razlog",as_index=False)["Minuta_iz_note"].sum().sort_values("Minuta_iz_note",ascending=False)
        st.markdown("### Ukupni zastoji po razlogu")
        _bar_chart(zbir.head(15),"Razlog","Minuta_iz_note","Ukupni zastoji po razlogu",True,"#ffb95f")
        prikazi_dark_dataframe(zbir)

        st.markdown("### Zastoji po procesu")
        proc=dfz.groupby("Proces",as_index=False)["Minuta_iz_note"].sum().sort_values("Minuta_iz_note",ascending=False)
        _bar_chart(proc,"Proces","Minuta_iz_note","Zastoji po procesu",False,"#4edea3")
        prikazi_dark_dataframe(proc)

        st.markdown("### Zastoji po projektu")
        proj=dfz.groupby("Projekat",as_index=False)["Minuta_iz_note"].sum().sort_values("Minuta_iz_note",ascending=False)
        _bar_chart(proj,"Projekat","Minuta_iz_note","Zastoji po projektu",False,"#adc6ff")
        prikazi_dark_dataframe(proj)

        st.markdown("### Zastoji po mašini")
        mas=dfz.groupby("Masina",as_index=False)["Minuta_iz_note"].sum().sort_values("Minuta_iz_note",ascending=False)
        _bar_chart(mas,"Masina","Minuta_iz_note","Zastoji po mašini",True,"#ffb4ab")
        prikazi_dark_dataframe(mas)

        st.markdown("### Detaljno po mašinama")
        _render_machine_detail_grid(dfz)

        st.markdown("### Top 3 uzroka po mašini i procesu")
        _render_top_causes_by_process(dfz)

        st.markdown("### Pareto uzroka zastoja")
        _pareto(dfz,"Razlog","Minuta_iz_note","Pareto uzroka zastoja")
        with st.expander("Prikaži sve stavke zastoja iz notes-a"):
            prikazi_dark_dataframe(dfz)

elif aktivna_sekcija == "KPI":
    st.subheader("📈 KPI — globalni")
    prikazi_gauge_dashboard("Ukupno filtrirano", df_ukupno_filter)
    st.markdown("### KPI po procesu")
    for proces in [p for p in KPI_PROCESI if p in df_ukupno_filter["Proces"].unique()]:
        with st.container():
            prikazi_gauge_dashboard(proces, df_ukupno_filter[df_ukupno_filter["Proces"] == proces])

elif aktivna_sekcija == "Grafički prikaz":
    st.subheader("Grafički prikaz")
    if df_filter.empty:
        st.info("Nema podataka za izabrane filtere.")
    else:
        trend = df_filter.groupby("Datum", as_index=False).agg({"Realizacija_STATOR": "sum", "Realizacija_ROTOR": "sum", "NOK_STATOR": "sum", "NOK_ROTOR": "sum", "Stops_min": "sum"})
        trend["Datum_txt"] = trend["Datum"].dt.strftime("%d.%m.%Y")
        trend["Realizacija"] = trend["Realizacija_STATOR"] + trend["Realizacija_ROTOR"]
        trend["NOK"] = trend["NOK_STATOR"] + trend["NOK_ROTOR"]
        fig = go.Figure()
        if tip_grafikona == "Linijski":
            fig.add_trace(go.Scatter(x=trend["Datum_txt"], y=trend["Realizacija"], mode="lines+markers", name="Realizacija"))
            fig.add_trace(go.Scatter(x=trend["Datum_txt"], y=trend["NOK"], mode="lines+markers", name="NOK"))
        elif tip_grafikona == "Površinski":
            fig.add_trace(go.Scatter(x=trend["Datum_txt"], y=trend["Realizacija"], fill="tozeroy", mode="lines", name="Realizacija"))
            fig.add_trace(go.Scatter(x=trend["Datum_txt"], y=trend["NOK"], fill="tozeroy", mode="lines", name="NOK"))
        else:
            fig.add_trace(go.Bar(x=trend["Datum_txt"], y=trend["Realizacija"], name="Realizacija"))
            fig.add_trace(go.Bar(x=trend["Datum_txt"], y=trend["NOK"], name="NOK"))
        st.plotly_chart(_dark_fig(fig, "Trend po danima"), use_container_width=True, config={"displayModeBar": False})
        prikazi_dark_dataframe(trend[["Datum_txt", "Realizacija", "NOK", "Stops_min"]].rename(columns={"Datum_txt": "Datum", "Stops_min": "Stops/min"}))

        _render_process_graphics(df_filter, "STAMPING", "AIDA / STAMPING REALIZACIJA")
        _render_process_graphics(df_filter, "ROTOR", "ROTOR LINIJE")
        _render_process_graphics(df_filter, "WELDING", "WELDING LINIJE")
        _render_process_graphics(df_filter, "DMC", "DMC LINIJE")

elif aktivna_sekcija == "SCRAP":
    st.subheader("SCRAP")
    if df_ukupno_filter.empty:
        st.info("Nema podataka za SCRAP.")
    else:
        scrap = df_ukupno_filter.groupby("Proces", as_index=False).agg({"OK_STATOR": "sum", "OK_ROTOR": "sum", "NOK_STATOR": "sum", "NOK_ROTOR": "sum"})
        scrap["OK"] = scrap["OK_STATOR"] + scrap["OK_ROTOR"]
        scrap["NOK"] = scrap["NOK_STATOR"] + scrap["NOK_ROTOR"]
        scrap["Scrap %"] = scrap.apply(lambda r: procenat(r["NOK"], r["OK"] + r["NOK"]), axis=1)
        fig = go.Figure(go.Bar(x=scrap["Proces"], y=scrap["NOK"], text=scrap["Scrap %"].apply(_fmt_pct_local), textposition="outside", marker_color="#ffb4ab"))
        st.plotly_chart(_dark_fig(fig, "NOK po procesu"), use_container_width=True, config={"displayModeBar": False})
        prikazi_dark_dataframe(scrap[["Proces", "OK", "NOK", "Scrap %"]])
        _render_scrap_sections(df_ukupno_filter)

elif aktivna_sekcija == "Top uzroci po mašini":
    st.subheader("Top uzroci po mašini")
    if df_zastoji_filter.empty and df_nok_filter.empty:
        st.info("Nema notes podataka za izabrane filtere.")
    else:
        masine = sorted(set(df_filter["Masina"].dropna().unique()))
        for masina in masine:
            col1, col2 = st.columns(2)
            with col1:
                z = df_zastoji_filter[df_zastoji_filter["Masina"] == masina] if not df_zastoji_filter.empty else pd.DataFrame()
                if not z.empty:
                    topz = z.groupby("Razlog", as_index=False)["Minuta_iz_note"].sum().sort_values("Minuta_iz_note", ascending=False).head(3)
                    st.markdown(f"<div class='neo-card'><div class='machine-title'>⏱️ {masina} — zastoji</div>", unsafe_allow_html=True)
                    for _, r in topz.iterrows():
                        st.markdown(f"<div class='reason-sub'>{_html_escape(r['Razlog'])}: <b>{_fmt_num(r['Minuta_iz_note'])} min</b></div>", unsafe_allow_html=True)
                    st.markdown("</div>", unsafe_allow_html=True)
            with col2:
                n = df_nok_filter[df_nok_filter["Masina"] == masina] if not df_nok_filter.empty else pd.DataFrame()
                if not n.empty:
                    topn = n.groupby("Razlog", as_index=False)["Komada_iz_note"].sum().sort_values("Komada_iz_note", ascending=False).head(3)
                    st.markdown(f"<div class='neo-card neo-card-purple'><div class='machine-title'>❌ {masina} — NOK</div>", unsafe_allow_html=True)
                    for _, r in topn.iterrows():
                        st.markdown(f"<div class='reason-sub'>{_html_escape(r['Razlog'])}: <b>{_fmt_num(r['Komada_iz_note'])} kom</b></div>", unsafe_allow_html=True)
                    st.markdown("</div>", unsafe_allow_html=True)

elif aktivna_sekcija == "Tabele":
    st.subheader("Tabele")
    st.markdown("### Dnevna tabela")
    dnevna = df_filter.copy()
    if "Datum" in dnevna.columns:
        dnevna["Datum"] = dnevna["Datum"].dt.strftime("%d.%m.%Y")
    prikazi_dark_dataframe(dnevna)

    st.markdown("### NOK notes tabela")
    nok_t = df_nok_filter.copy()
    if not nok_t.empty and "Datum" in nok_t.columns:
        nok_t["Datum"] = nok_t["Datum"].dt.strftime("%d.%m.%Y")
    prikazi_dark_dataframe(nok_t)

    st.markdown("### Zastoji notes tabela")
    z_t = df_zastoji_filter.copy()
    if not z_t.empty and "Datum" in z_t.columns:
        z_t["Datum"] = z_t["Datum"].dt.strftime("%d.%m.%Y")
    prikazi_dark_dataframe(z_t)
